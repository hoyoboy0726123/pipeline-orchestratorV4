"""
Pipeline 狀態機主引擎。

流程：
  START → 逐步執行 → LLM 驗證 → 通過則下一步
                                 → 失敗且有重試次數 → 自動重試
                                 → 失敗且重試耗盡  → 暫停 + Telegram inline keyboard
  用戶按 [重試 / 跳過 / 中止] → resume_pipeline() 繼續或結束

Telegram 通知時機：
  - 步驟失敗需人為決策 → 詢問訊息 + inline keyboard
  - Pipeline 全部完成 / 中止 → 結果摘要
"""
import asyncio
import logging
import os
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional

from telegram import Bot, InlineKeyboardButton, InlineKeyboardMarkup

from config import TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID
from .models import PipelineConfig
from .store import PipelineRun, StepResult, get_store
from .logger import create_run_logger, resume_run_logger
from .executor import execute_step, execute_step_with_skill
from .validator import validate_step, validate_step_with_skill, ValidationResult


# ── Abort flags（in-memory）────────────────────────────────────────────────────
_abort_flags: set[str] = set()

# ── Running task tracking（for immediate cancel）──────────────────────────────
_running_tasks: dict[str, asyncio.Task] = {}


def register_task(run_id: str, task: asyncio.Task):
    _running_tasks[run_id] = task


def unregister_task(run_id: str):
    _running_tasks.pop(run_id, None)


def request_abort(run_id: str):
    """前端/API 呼叫：標記此 run 需要中止"""
    _abort_flags.add(run_id)


async def force_abort(run_id: str):
    """立即中止：kill 子進程 + 標記 computer_use abort + cancel asyncio task + 更新狀態"""
    from .executor import kill_run_processes
    from .computer_use import request_abort as _cu_abort
    _abort_flags.add(run_id)
    # 1. 立即 kill 所有子進程
    kill_run_processes(run_id)
    # 1a. 通知 computer_use 引擎中止（它跑在 executor thread 裡，kill 不到）
    _cu_abort(run_id)
    # 2. Cancel asyncio task
    task = _running_tasks.pop(run_id, None)
    if task and not task.done():
        task.cancel()
        try:
            await task
        except (asyncio.CancelledError, Exception):
            pass
    # 3. 更新 run 狀態
    store = get_store()
    run = store.load(run_id)
    if run and run.status in ("running", "awaiting_human"):
        run.status = "aborted"
        run.ended_at = datetime.now().isoformat()
        store.save(run)
        logger = logging.getLogger(f"pipeline.{run_id}")
        logger.info("⛔ Pipeline 被立即中止（force abort）")
        try:
            config = PipelineConfig.from_dict(run.config_dict)
            await _notify_final(run, config)
        except Exception:
            pass
    clear_abort(run_id)


def is_abort_requested(run_id: str) -> bool:
    return run_id in _abort_flags


def clear_abort(run_id: str):
    _abort_flags.discard(run_id)


# ── Telegram helpers ─────────────────────────────────────────────────────────

def _decision_keyboard(run_id: str) -> InlineKeyboardMarkup:
    # 步驟失敗時的決策鍵盤。
    # 💡 截圖按鈕：失敗時使用者可能人不在電腦前，加截圖按鈕讓遠端也能先看畫面再決策
    #    （不分節點類型 — skill 腳本 / 桌面自動化 / shell 失敗都可能需要看現場）
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("🔄 重試", callback_data=f"pipe_retry:{run_id}"),
            InlineKeyboardButton("💬 補充指示", callback_data=f"pipe_hint:{run_id}"),
        ],
        [
            InlineKeyboardButton("📸 截圖", callback_data=f"pipe_screenshot:{run_id}"),
            InlineKeyboardButton("📋 查看 Log", callback_data=f"pipe_log:{run_id}"),
        ],
        [
            InlineKeyboardButton("🛑 中止", callback_data=f"pipe_abort:{run_id}"),
        ],
    ])


def _confirm_keyboard(run_id: str, screenshot: bool = False, allow_hint: bool = False,
                      preview_enabled: bool = False) -> InlineKeyboardMarkup:
    # 人工確認節點的按鈕。allow_hint 只在「上一個可執行節點是 AI 技能（skill_mode）」時 True。
    # preview_enabled 只在 step.preview_prev_output=True 時 True（=自動預覽有啟用才給 HQ 選項）
    top = [InlineKeyboardButton("✅ 繼續執行", callback_data=f"pipe_continue:{run_id}")]
    if allow_hint:
        top.append(InlineKeyboardButton("💬 補充指示", callback_data=f"pipe_hint:{run_id}"))
    top.append(InlineKeyboardButton("🛑 中止", callback_data=f"pipe_abort:{run_id}"))
    rows = [
        top,
        # 「📎 上一步輸出 / 📂 任一步輸出」永遠存在；點下去由 backend 判斷有沒有檔案、回應使用者
        # 跟 send_prev_output 自動傳送無關（自動傳是抵達節點當下推一份；按鈕是隨時要重抓用）
        [InlineKeyboardButton("📎 上一步輸出", callback_data=f"pipe_prev_output:{run_id}"),
         InlineKeyboardButton("📂 任一步輸出", callback_data=f"pipe_select_step:{run_id}")],
        [InlineKeyboardButton("📋 查看 Log", callback_data=f"pipe_log:{run_id}")],
    ]
    if screenshot:
        rows[2].append(InlineKeyboardButton("📸 截圖", callback_data=f"pipe_screenshot:{run_id}"))
    # HQ 預覽：B1 自動預覽只抽文字（docx/pptx 品質 40%），點此按鈕改用 LibreOffice
    # 轉 PDF → render，版式 ~80-90% 還原。要 5-10s，所以不自動跑；使用者按了才跑。
    if preview_enabled:
        rows.append([InlineKeyboardButton("🎨 原版式預覽（LibreOffice）",
                                          callback_data=f"pipe_preview_hq:{run_id}")])
    return InlineKeyboardMarkup(rows)


def _ask_user_keyboard(run_id: str, options: list) -> InlineKeyboardMarkup:
    """
    ask_user 問題的 Telegram 鍵盤。
    - 有 options → 每個選項一個按鈕（一行最多 2 個）+ 自由輸入 + 中止
    - 無 options → 只有「自由輸入」和「中止」
    """
    rows: list[list[InlineKeyboardButton]] = []
    if options:
        # callback 長度上限 64 bytes，用索引傳遞
        row: list[InlineKeyboardButton] = []
        for i, opt in enumerate(options):
            label = str(opt)
            if len(label) > 30:
                label = label[:27] + "…"
            row.append(InlineKeyboardButton(label, callback_data=f"pipe_answer:{run_id}:{i}"))
            if len(row) == 2:
                rows.append(row)
                row = []
        if row:
            rows.append(row)
    rows.append([
        InlineKeyboardButton("✍ 自由輸入", callback_data=f"pipe_answer_free:{run_id}"),
        InlineKeyboardButton("🛑 中止", callback_data=f"pipe_abort:{run_id}"),
    ])
    return InlineKeyboardMarkup(rows)


async def _send_ask_user_notification(run, question: str, options: list, context: str, step_name: str):
    """Skill agent 呼叫 ask_user 時發送 Telegram 通知。"""
    import html as _html
    total = len(run.config_dict.get("steps", [])) if run.config_dict else 0
    step_num = run.current_step + 1
    lines = [
        "❓ <b>AI 技能請求人工協助</b>",
        "",
        f"📋 {run.pipeline_name}",
        f"📍 步驟 {step_num}/{total}：<b>{_html.escape(step_name)}</b>",
        "",
        f"<b>問題</b>：{_html.escape(question)}",
    ]
    if context:
        lines.append(f"\n<b>背景</b>：{_html.escape(context)}")
    if options:
        lines.append("\n請從下方選項選擇，或點「自由輸入」回答。")
    else:
        lines.append("\n請點「自由輸入」並傳送文字回答。")
    await _tg_send(run.telegram_chat_id, "\n".join(lines),
                   _ask_user_keyboard(run.run_id, options))


def _is_valid_tg_token(token: str) -> bool:
    """檢查 Telegram Bot Token 格式是否正確（數字:字母混合）"""
    if not token or ":" not in token:
        return False
    parts = token.split(":", 1)
    return parts[0].isdigit() and len(parts[1]) > 10


def _get_tg_token() -> str:
    """取得 Telegram Bot Token（優先用 settings UI 設定，fallback 到 env）"""
    logger = logging.getLogger("pipeline")
    try:
        from settings import get_settings
        token = get_settings().get("telegram_bot_token", "")
        if token and _is_valid_tg_token(token):
            return token
        elif token:
            logger.warning(f"[Telegram] settings 中的 token 格式不正確（'{token[:15]}...'），改用 .env")
    except Exception:
        pass
    if TELEGRAM_BOT_TOKEN:
        logger.debug(f"[Telegram] 使用 .env 的 TELEGRAM_BOT_TOKEN")
    return TELEGRAM_BOT_TOKEN


def _get_tg_chat_id() -> int:
    """取得 Telegram Chat ID（優先 settings UI，fallback 到 env）"""
    logger = logging.getLogger("pipeline")
    try:
        from settings import get_settings
        cid = get_settings().get("telegram_chat_id", "")
        if cid:
            return int(cid)
    except Exception:
        pass
    # fallback 到 .env
    if TELEGRAM_CHAT_ID:
        try:
            logger.debug(f"[Telegram] 使用 .env 的 TELEGRAM_CHAT_ID")
            return int(TELEGRAM_CHAT_ID)
        except ValueError:
            logger.warning(f"[Telegram] .env TELEGRAM_CHAT_ID 格式不正確：{TELEGRAM_CHAT_ID}")
    return 0


async def _tg_send(chat_id: int, text: str, reply_markup=None):
    """發送 Telegram 訊息（錯誤靜默記錄，不拋出）"""
    logger = logging.getLogger("pipeline")
    token = _get_tg_token()
    # 如果沒傳 chat_id，嘗試從 settings 取得
    if not chat_id:
        chat_id = _get_tg_chat_id()
    if not chat_id or not token:
        logger.warning(f"[Telegram] 跳過發送：chat_id={chat_id}, token={'有' if token else '無'}")
        return
    logger.info(f"[Telegram] 發送訊息到 chat_id={chat_id}（token={token[:15]}...）")
    try:
        # 用 async with 取代手動 bot.close() — 後者實際是 TG API 的 `close` method
        # （TG 文件警告：前 10 分鐘必回 429、嚴格 rate-limit、不該在 bot code 呼叫）
        # async with 走 shutdown() 路徑、只關 httpx 連線、不打 TG API
        async with Bot(token=token) as bot:
            await bot.send_message(
                chat_id=chat_id,
                text=text,
                parse_mode="HTML",
                reply_markup=reply_markup,
            )
        logger.info(f"[Telegram] ✅ 發送成功")
    except Exception as e:
        logger.error(f"[Telegram] ❌ 發送失敗：{e}")


# 送 TG photo 的壓縮參數：一律壓縮（不看原檔大小），讓每張 traffic 一致、傳輸時間接近
# → 避免「大的壓了變小、小的沒壓還是大」的不對稱上傳時間造成誤判 timeout + 重複訊息
# 長寬上限：1920（TG 本來就會壓到 ~1280 顯示，1920 已經足夠清楚，肉眼看不出差）
_TG_PHOTO_MAX_DIM = 1920
_TG_PHOTO_JPEG_Q  = 85


def _compress_for_tg(src_path: str) -> str:
    """一律轉 JPEG + 縮邊到 _TG_PHOTO_MAX_DIM。回傳新產生的 _compressed.jpg 路徑。
    Pillow 缺席 / 讀圖失敗 → 回原路徑當 fallback。
    為什麼不看門檻：上次 bug 就是 mon1 沒壓（大）+ mon2 壓了（小）→ mon1 上傳慢 120s 超時誤判。
    統一都壓就沒這問題，而且 TG 顯示時本來就壓到 ~1280，我們先壓 1920 剛剛好。
    """
    logger = logging.getLogger("pipeline")
    try:
        from pathlib import Path as _P
        src = _P(src_path)
        if not src.exists():
            return src_path
        orig_size = src.stat().st_size
        try:
            from PIL import Image
        except Exception:
            logger.warning(f"[Telegram] Pillow 未安裝，照原圖送 {src.name}（{orig_size/1024/1024:.1f}MB）")
            return src_path
        im = Image.open(src)
        if im.mode not in ("RGB", "L"):
            im = im.convert("RGB")
        w, h = im.size
        if max(w, h) > _TG_PHOTO_MAX_DIM:
            scale = _TG_PHOTO_MAX_DIM / max(w, h)
            im = im.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
        out = src.with_suffix("")
        out = out.with_name(out.name + "_compressed.jpg")
        im.save(out, "JPEG", quality=_TG_PHOTO_JPEG_Q, optimize=True)
        new_size = out.stat().st_size
        logger.info(
            f"[Telegram] 壓縮 {src.name}：{orig_size/1024/1024:.2f}MB ({w}×{h}) "
            f"→ {out.name}：{new_size/1024/1024:.2f}MB {im.size}"
        )
        return str(out)
    except Exception as e:
        logger.warning(f"[Telegram] 壓縮失敗（照原圖送）：{e}")
        return src_path


async def _tg_send_photos(chat_id: int, paths: list[str], caption_prefix: str = ""):
    """批次送截圖。每張 caption「螢幕 k/n」。
    流程：超過 4.5MB → Pillow 壓到 JPEG+1920 邊再送；send_photo 還是失敗 → 退 send_document。
    每張分開 try/except，一張壞不會拖垮整批。
    """
    logger = logging.getLogger("pipeline")
    if not paths:
        return
    token = _get_tg_token()
    if not chat_id:
        chat_id = _get_tg_chat_id()
    if not chat_id or not token:
        logger.warning(f"[Telegram] 批次截圖跳過：chat_id={chat_id} token={'有' if token else '無'}")
        return
    logger.info(f"[Telegram] 批次送 {len(paths)} 張截圖 → chat_id={chat_id}")

    # send 每張的硬超時：避免 send_photo hang 把 poll loop 整個卡死
    _PHOTO_TIMEOUT_S = 120

    # 重複訊息 root cause：timeout / network error 時 Python 以為失敗，但 TG 其實已收到，
    # 我們又送了一次 document → 使用者收到 2 份同內容。
    # 修法：只在「確定 TG 拒收這個檔」（BadRequest，例如格式/尺寸錯）時才 fallback document；
    #       timeout / NetworkError / TimedOut 都視為「很可能已送達」不重送。
    from telegram.error import BadRequest as _TgBadRequest  # noqa: WPS433 (局部 import 沒關係)

    async def _send_one(bot, send_path: str, cap: str, i: int, total: int) -> bool:
        """送單張。回傳 True=已送達（或近乎送達），False=徹底失敗。"""
        # 1) 先試 send_photo
        try:
            with open(send_path, "rb") as fh:
                await asyncio.wait_for(
                    bot.send_photo(chat_id=chat_id, photo=fh, caption=cap or None),
                    timeout=_PHOTO_TIMEOUT_S,
                )
            logger.info(f"[Telegram]   ✓ 送出截圖 {i}/{total}")
            return True
        except _TgBadRequest as e:
            # 檔案格式 / 尺寸被 TG 拒收 — 真的壞了，退 document 才有意義
            logger.warning(f"[Telegram]   photo {i}/{total} TG 拒收（BadRequest），退 document：{e}")
        except asyncio.TimeoutError:
            logger.warning(
                f"[Telegram]   photo {i}/{total} 超過 {_PHOTO_TIMEOUT_S}s 沒回 ack，"
                f"TG 可能已收到（不重送 document 避免重複）"
            )
            return True
        except Exception as e:
            # network / httpx timeout / retry-after 等 — TG 是否收到不確定；
            # 為避免重複訊息，一律視為「可能已送達」不 fallback（之前 case 就是這裡誤判）
            logger.warning(
                f"[Telegram]   photo {i}/{total} 送出時出例外（{type(e).__name__}: {e}）— "
                f"TG 可能已收到，不重送 document 以避免重複"
            )
            return True
        # 2) Fallback：send_document（只在真的 BadRequest 時才走）
        try:
            with open(send_path, "rb") as fh:
                await asyncio.wait_for(
                    bot.send_document(chat_id=chat_id, document=fh, caption=cap or None),
                    timeout=_PHOTO_TIMEOUT_S,
                )
            logger.info(f"[Telegram]   ✓ 以 document 形式送出截圖 {i}/{total}")
            return True
        except asyncio.TimeoutError:
            logger.warning(f"[Telegram]   document {i}/{total} 超時但 TG 可能已收到")
            return True
        except Exception as e2:
            logger.error(f"[Telegram]   ✗ 截圖 {i}/{total} 送出徹底失敗：{type(e2).__name__}: {e2}")
        return False

    try:
        # 用 async with：避免手動 bot.close()（那是 TG API、嚴格 rate-limit）
        async with Bot(token=token) as bot:
            total = len(paths)
            for i, p in enumerate(paths, start=1):
                cap = caption_prefix + (f"（螢幕 {i}/{total}）" if total > 1 else "")
                send_path = _compress_for_tg(p)
                # 送每張前檢查檔案有沒有正常產生（有時 take_screenshots 該路徑被清掉或空檔）
                try:
                    sz = os.path.getsize(send_path)
                except Exception as e:
                    logger.error(f"[Telegram]   截圖 {i}/{total} 檔案讀取失敗（{e}）→ 跳過")
                    continue
                if sz <= 0:
                    logger.error(f"[Telegram]   截圖 {i}/{total} 檔案 0 bytes → 跳過")
                    continue
                ok = await _send_one(bot, send_path, cap, i, total)
                # 送成功 → 清掉磁碟（截圖 TG 上已經有，本地不留以免每跑一次就堆積 GB 級 PNG）
                # 送失敗 → 保留讓使用者/開發者事後回看或手動重送
                if ok:
                    for cleanup in {p, send_path}:  # set 去重：單螢幕沒壓縮時兩者同檔
                        try:
                            if os.path.exists(cleanup):
                                os.unlink(cleanup)
                        except Exception as _e:
                            logger.warning(f"[Telegram]   清理截圖 {cleanup} 失敗：{_e}")
    except Exception as e:
        logger.error(f"[Telegram] 截圖批次送出異常：{e}")


# Telegram 文件大小上限（一般 bot；自架 local server 可放寬到 2GB）
_TG_DOC_MAX_BYTES = 50 * 1024 * 1024


def _workflow_output_dir(workflow_name: str):
    """回傳 ai_output/<workflow_name>/ 的絕對路徑（不存在也回，由呼叫端決定怎麼處理）。"""
    from pathlib import Path as _P
    if not workflow_name:
        return None
    proj_root = _P(__file__).parent.parent.parent.absolute()
    return proj_root / "ai_output" / workflow_name


# 用來判斷哪些檔案是「真正的步驟產出」、哪些是雜訊（log / preview / 內部 DB 檔）
_OUTPUT_SKIP_PREFIXES = ("screenshot_",)
_OUTPUT_SKIP_SUFFIXES = ("_preview.png", "_compressed.jpg", "_libre.pdf", "_unsupported.png")
_OUTPUT_SKIP_EXTS = {".log"}
_OUTPUT_SKIP_NAMES = ("pipeline_settings.json", "pipeline.db", "pipeline.db-shm", "pipeline.db-wal")


def _is_output_candidate(path) -> bool:
    """檔名是否該被視為「步驟可能產出的檔案」。過濾掉系統雜訊。"""
    n = path.name
    if n.startswith(_OUTPUT_SKIP_PREFIXES):
        return False
    if any(n.endswith(suf) for suf in _OUTPUT_SKIP_SUFFIXES):
        return False
    if path.suffix.lower() in _OUTPUT_SKIP_EXTS:
        return False
    if n in _OUTPUT_SKIP_NAMES:
        return False
    return True


def _snapshot_workflow_dir(workflow_name: str) -> dict:
    """掃 ai_output/<workflow_name>/ 取每個檔的 mtime（給步驟前後比對用）。
    回 {str(absolute_path): mtime}。失敗回空 dict。"""
    out: dict = {}
    wf = _workflow_output_dir(workflow_name)
    if not wf or not wf.exists() or not wf.is_dir():
        return out
    try:
        for f in wf.rglob("*"):
            if f.is_file() and _is_output_candidate(f):
                try:
                    out[str(f.absolute())] = f.stat().st_mtime
                except OSError:
                    pass
    except Exception:
        pass
    return out


def _diff_snapshot_pick_main(before: dict, workflow_name: str):
    """比對 before（_snapshot_workflow_dir 結果）跟現在的狀態、找出本步驟新增/修改的檔，
    挑「主要產出」回傳（絕對路徑字串），沒有變化回 None。

    挑選邏輯：
      1. 先看新增檔（before 沒有的） — 比修改現有檔更可能是「最終產出」
      2. 都沒新增、看修改的（mtime 變新）
      3. 多個候選時，偏好「報告類副檔名」（.docx/.pdf/.xlsx/.md/.csv/.html/.pptx/.json/.txt）
         排在前面、再看 mtime 最新的
    """
    wf = _workflow_output_dir(workflow_name)
    if not wf or not wf.exists():
        return None
    # 取現在快照
    after = _snapshot_workflow_dir(workflow_name)

    new_files = [p for p in after.keys() if p not in before]
    modified = [p for p in after.keys() if p in before and after[p] > before[p]]
    candidates = new_files if new_files else modified
    if not candidates:
        return None

    from pathlib import Path as _P
    # 報告類副檔名優先（排前面）
    report_exts = {".docx", ".pdf", ".xlsx", ".xls", ".pptx", ".md", ".csv",
                   ".html", ".htm", ".json", ".txt", ".png", ".jpg", ".jpeg"}

    def sort_key(p_str: str):
        p = _P(p_str)
        is_report = 0 if p.suffix.lower() in report_exts else 1
        # 用負 mtime 讓「最新」排前面
        return (is_report, -after.get(p_str, 0))

    candidates.sort(key=sort_key)
    return candidates[0]


def _latest_workflow_output_file(workflow_name: str):
    """掃 ai_output/<workflow_name>/ 拿最新一個非雜訊檔（圖檔 / log / preview / 內部 db 檔過濾掉）。
    給「skill 節點 / 沒明確 output.path」這種「實際有產檔但 step 沒記錄」的場景兜底。
    """
    from pathlib import Path as _P
    if not workflow_name:
        return None
    proj_root = _P(__file__).parent.parent.parent.absolute()
    wf_dir = proj_root / "ai_output" / workflow_name
    if not wf_dir.exists() or not wf_dir.is_dir():
        return None
    skip_prefixes = ("screenshot_",)
    skip_suffixes = ("_preview.png", "_compressed.jpg", "_libre.pdf", "_unsupported.png")
    skip_exts = {".log"}
    skip_names = ("pipeline_settings.json", "pipeline.db", "pipeline.db-shm", "pipeline.db-wal")
    candidates = []
    for f in wf_dir.iterdir():
        if not f.is_file():
            continue
        n = f.name
        if n.startswith(skip_prefixes):
            continue
        if any(n.endswith(suf) for suf in skip_suffixes):
            continue
        if f.suffix.lower() in skip_exts:
            continue
        if n in skip_names:
            continue
        candidates.append((f.stat().st_mtime, f))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def _step_default_output_path(step, workflow_name: str) -> Optional[str]:
    """如果 step 沒明確設 output.path、回傳 runner 會自動 default 的路徑。
    V4 的 skill / script / human_confirm / visual_validation / computer_use 節點都沒 default
    rule（runner 不會自動代填 output.path），統一回 None。
    """
    return None


def _resolve_step_output_for_tg(
    step, *, workflow_name: str = "", logger=None, step_result=None,
) -> tuple[Optional[str], Optional[str], Optional[str]]:
    """檢查指定 step 的輸出檔、回傳能傳給 Telegram 的檔案資料。

    解析優先順序：
      1. step_result.actual_output_path（runner snapshot diff 算出來的）
      2. step.output.path 明確設定 → 用它
      3. 沒設、但節點類型有 default rule → 算出 default 路徑（V4 沒 default 規則 → None）
      4. 都沒有 → workflow dir 最新檔（兜底）
      5. 都沒有 → None

    回傳 (file_path, display_name, error_msg)：
      - 三者皆 None：完全沒輸出可傳
      - file_path / display_name 給值、error_msg=None：可以傳
      - error_msg 有值：能解析但有狀況（不存在 / 太大），傳給使用者看

    資料夾：自動 zip 後給；zip 寫到系統 temp、呼叫端用完要 unlink。
    """
    from pathlib import Path as _P
    import logging as _log

    log_fn = (logger or _log.getLogger("pipeline")).warning

    if not step:
        return None, None, None

    # ── Step 1：抓明確或 default 的目標路徑 ──────────────────────
    target_str: Optional[str] = None
    is_explicit = False
    # 最高優先：StepResult.actual_output_path（runner 在執行時 snapshot diff 算出來的）
    # 這個對「沒設 output.path 的 skill 節點」特別重要 — 不會跟其他步驟搶 latest 檔
    actual_path = getattr(step_result, "actual_output_path", "") if step_result else ""
    if actual_path:
        target_str = actual_path
        is_explicit = True  # 視為「指定路徑」
    elif getattr(step, "output", None) and step.output.path:
        target_str = str(step.output.path)
        is_explicit = True
    else:
        target_str = _step_default_output_path(step, workflow_name)

    # Skill / Script 沒 actual_output_path、沒明確 output.path 也沒 default rule，
    # 通常表示：(a) 該步沒實際寫檔（如純 stdout）；或 (b) 是舊 run（升級前沒記錄 actual_output_path）
    # 對 (b) 兜底掃 workflow dir 最新檔（不完美但聊勝於無）
    # human_confirm / visual_validation / computer_use 不寫檔、不適用。
    if not target_str and workflow_name:
        could_produce = bool(
            getattr(step, "skill_mode", False)
            or (not getattr(step, "human_confirm", False)
                and not getattr(step, "visual_validation", False)
                and not getattr(step, "computer_use", False)
                and getattr(step, "batch", ""))  # script 節點：有 batch 表示會跑
        )
        if could_produce:
            latest = _latest_workflow_output_file(workflow_name)
            if latest:
                target_str = str(latest)
                if logger:
                    logger.info(f"[_resolve_step_output_for_tg] step={step.name} 沒設 output、"
                                f"fallback 到 workflow dir 最新檔：{latest.name}")

    if not target_str:
        return None, None, None

    p = _P(target_str).expanduser()
    if not p.is_absolute():
        p = _P(__file__).parent.parent.parent.absolute() / p

    if not p.exists():
        return None, None, f"輸出檔案不存在：{p}"

    if p.is_file():
        size = p.stat().st_size
        if size > _TG_DOC_MAX_BYTES:
            return None, None, (f"檔案太大（{size/1024/1024:.1f} MB > 50 MB Telegram 上限）。"
                                f"請去 host 取：{p}")
        return str(p), p.name, None

    if p.is_dir():
        # 整個資料夾打包成 zip 送
        import tempfile, zipfile
        try:
            tmp_zip = tempfile.NamedTemporaryFile(suffix=".zip", delete=False).name
            with zipfile.ZipFile(tmp_zip, "w", zipfile.ZIP_DEFLATED) as zf:
                for f in p.rglob("*"):
                    if f.is_file():
                        # 用 dir 自身為 root；解壓會看到原本資料夾結構
                        zf.write(f, f.relative_to(p.parent))
            zsize = _P(tmp_zip).stat().st_size
            if zsize > _TG_DOC_MAX_BYTES:
                try: os.unlink(tmp_zip)
                except Exception: pass
                return None, None, (f"資料夾打包後 {zsize/1024/1024:.1f} MB、超過 Telegram 上限。"
                                    f"請去 host 取：{p}")
            return tmp_zip, p.name + ".zip", None
        except Exception as e:
            log_fn(f"[_resolve_step_output_for_tg] zip 失敗：{e}")
            return None, None, f"資料夾打包失敗：{e}"

    return None, None, f"不認識的路徑類型：{p}"


async def _send_step_output_to_tg(
    chat_id: int, step, step_label: str = "", *,
    workflow_name: str = "", logger=None, step_result=None,
) -> tuple[bool, str]:
    """把 step 的 output 檔案（或 zip 化的資料夾）送到指定 chat_id。
    workflow_name 用來算 default output path（沒填 output.path 但節點類型有 default 時）。
    step_result：StepResult 物件（含 actual_output_path）— 多步 skill 共用 workflow dir 時不會搶錯檔。
    回傳 (ok, msg)：ok=True 時 msg 是 status 摘要；False 時 msg 是錯誤訊息（可直接給使用者）。
    """
    import logging as _log
    import tempfile as _tf
    log = logger or _log.getLogger("pipeline")
    file_path, display, err = _resolve_step_output_for_tg(
        step, workflow_name=workflow_name, logger=log, step_result=step_result,
    )

    if err:
        return False, err
    if not file_path:
        return False, "上一步沒設輸出檔（output.path），且該節點類型沒有 default fallback"

    # chat_id fallback：跟 _tg_send / _tg_send_photos 同邏輯。
    # 之前 auto-send 會失敗的就是這裡 — chat_id 0 時沒退到 settings 拿
    if not chat_id:
        chat_id = _get_tg_chat_id()
    token = _get_tg_token()
    if not chat_id or not token:
        return False, f"Telegram 設定不完整：chat_id={chat_id or '無'}, token={'有' if token else '無'}"

    try:
        # 用 async with：避免手動 bot.close()（那是 TG `close` API method、
        # 文件寫前 10 分鐘必回 429、嚴格 rate-limit、不該在 bot code 呼叫）。
        # 之前 user 報「每次手動點按鈕必出現速率限制警告」就是這個 bug。
        async with Bot(token=token) as bot:
            with open(file_path, "rb") as fp:
                await bot.send_document(
                    chat_id=chat_id,
                    document=fp,
                    filename=display,
                    caption=f"📎 {step_label or '上一步輸出'}：{display}" if step_label else None,
                )
        log.info(f"[Telegram] ✅ 已傳送 {display} 到 chat {chat_id}")
        return True, f"已傳送：{display}"
    except Exception as e:
        # Telegram rate limit (flood control) → 翻譯成易懂訊息
        es = str(e)
        if "Flood control" in es or "RetryAfter" in es or "Too Many Requests" in es:
            import re as _re
            m = _re.search(r"(\d+)\s*seconds", es)
            wait_s = int(m.group(1)) if m else 0
            log.warning(f"[Telegram] FloodWait：{es}")
            return False, (
                f"Telegram 速率限制（短時間內傳太多訊息了）。"
                f"{f'請等 {wait_s} 秒（約 {wait_s//60} 分 {wait_s%60} 秒）後再試。' if wait_s else '稍候幾分鐘再試。'}"
            )
        log.error(f"[Telegram] send_document 失敗：{e}")
        return False, f"Telegram 傳送失敗：{e}"
    finally:
        # _resolve_step_output_for_tg 對「資料夾」會 zip 到 system temp 下，這邊清掉
        try:
            if file_path.startswith(_tf.gettempdir()) and file_path.endswith(".zip"):
                os.unlink(file_path)
        except Exception:
            pass


def _find_prev_output_file(run, config) -> Optional[str]:
    """找上一個非 human_confirm 步驟的輸出檔案。人工確認節點「附檔案預覽」用。
    策略：
      1. 往前找第一個 step.output.path 有設、且檔案存在 → 回傳它
      2. 若都沒設 output.path（或設了但檔案不存在）→ 退到預設資料夾 ai_output/<workflow>/
         抓最近修改時間最新的檔案（排除目錄、截圖、log、.json 設定等雜訊）
      3. 都找不到 → None
    """
    try:
        from pathlib import Path as _P
        import time as _t

        # 策略 1：看 step.output.path
        idx = run.current_step - 1
        while idx >= 0:
            st = config.steps[idx]
            if st.human_confirm:
                idx -= 1
                continue
            if st.output and st.output.path:
                p = _P(st.output.path).expanduser()
                if p.exists() and p.is_file():
                    return str(p)
            idx -= 1

        # 策略 2：預設目錄最新檔
        # 規則跟 main.py / take_screenshots 一致：ai_output/<pipeline_name>/
        proj_root = _P(__file__).parent.parent.parent.absolute()
        wf_dir = proj_root / "ai_output" / run.pipeline_name
        if not wf_dir.exists() or not wf_dir.is_dir():
            return None
        # 過濾規則：
        #   排除資料夾
        #   排除我們自己產的截圖（screenshot_*.png / _preview.png / _compressed.jpg）
        #   排除 log / 內部 JSON
        skip_prefixes = ("screenshot_",)
        skip_suffixes = ("_preview.png", "_compressed.jpg", "_libre.pdf", "_unsupported.png")
        skip_exts = {".log"}
        candidates = []
        for f in wf_dir.iterdir():
            if not f.is_file():
                continue
            name = f.name
            if name.startswith(skip_prefixes):
                continue
            if any(name.endswith(suf) for suf in skip_suffixes):
                continue
            if f.suffix.lower() in skip_exts:
                continue
            # 也排除 pipeline_settings / recipes / runs 等已知內部檔（以防掃到 OUTPUT_BASE）
            if name in ("pipeline_settings.json", "pipeline.db", "pipeline.db-shm", "pipeline.db-wal"):
                continue
            candidates.append((f.stat().st_mtime, f))
        if not candidates:
            return None
        candidates.sort(key=lambda x: x[0], reverse=True)
        latest = candidates[0][1]
        age = _t.time() - candidates[0][0]
        # 太舊的（>7 天）可能只是殘留，警告一下但還是回傳
        if age > 7 * 86400:
            import logging as _lg
            _lg.getLogger("pipeline").info(
                f"[preview] 退回找預設目錄最新檔，但最新檔已 {age/86400:.1f} 天沒更新，可能是殘留：{latest}"
            )
        return str(latest)
    except Exception:
        pass
    return None


def take_screenshots(pipeline_name: str, step_name: str) -> list[str]:
    """逐螢幕截圖（1 螢幕 → 1 張、2 螢幕 → 2 張）。回傳檔案路徑 list。
    mss 的 sct.monitors[0] 是「所有螢幕拼成的虛擬桌面」、monitors[1..N] 是每台實體螢幕。
    Telegram 看起來更直覺（不會因為多螢幕被壓成超寬一張），所以直接逐螢幕抓。
    """
    import time as _t
    from pathlib import Path as _P
    logger = logging.getLogger("pipeline")
    results: list[str] = []
    try:
        import mss as _mss
        _PROJ_ROOT = _P(__file__).parent.parent.parent.absolute()
        ss_dir = _PROJ_ROOT / "ai_output" / pipeline_name
        ss_dir.mkdir(parents=True, exist_ok=True)
        ts = _t.strftime('%Y%m%d_%H%M%S')
        with _mss.mss() as sct:
            monitors = sct.monitors[1:] or sct.monitors  # 單螢幕系統 monitors[1:] 可能空，退回全部
            for idx, mon in enumerate(monitors, start=1):
                tag = f"mon{idx}" if len(monitors) > 1 else "full"
                ss_path = ss_dir / f"screenshot_{step_name}_{ts}_{tag}.png"
                try:
                    img = sct.grab(mon)
                    from mss.tools import to_png as _to_png
                    _to_png(img.rgb, img.size, output=str(ss_path))
                except Exception as e:
                    logger.warning(f"[{step_name}] 螢幕 {idx} 截圖失敗（略過）：{e}")
                    continue
                # 確認檔案真的產生且不是 0 bytes（to_png 偶爾會沉默失敗）
                if not ss_path.exists():
                    logger.warning(f"[{step_name}] 螢幕 {idx} 截圖檔未產生：{ss_path}")
                    continue
                fsize = ss_path.stat().st_size
                if fsize <= 0:
                    logger.warning(f"[{step_name}] 螢幕 {idx} 截圖 0 bytes，刪除並略過：{ss_path}")
                    try:
                        ss_path.unlink()
                    except Exception:
                        pass
                    continue
                logger.info(f"[{step_name}]   ✓ 螢幕 {idx} 截圖 {fsize/1024:.0f} KB → {ss_path.name}")
                results.append(str(ss_path))
        if results:
            logger.info(f"[{step_name}] 📸 截圖 {len(results)}/{len(monitors)} 張已儲存")
        else:
            logger.warning(f"[{step_name}] 截圖失敗：沒有檔案產生")
    except Exception as e:
        logger.warning(f"[{step_name}] 截圖失敗：{e}")
    return results


def take_screenshot(pipeline_name: str, step_name: str) -> Optional[str]:
    """舊接口保留：回傳第一張截圖（向後相容，如有其他呼叫者）"""
    paths = take_screenshots(pipeline_name, step_name)
    return paths[0] if paths else None


async def _notify_failure(run: PipelineRun, val: ValidationResult, step_name: str):
    """詢問用戶如何處理失敗步驟"""
    step_num = run.current_step + 1
    total = len(PipelineConfig.from_dict(run.config_dict).steps)
    text = (
        f"⚠️ <b>Pipeline 需要決策</b>\n\n"
        f"📋 {run.pipeline_name}\n"
        f"📍 步驟 {step_num}/{total}：<b>{step_name}</b>\n\n"
        f"🔴 {val.reason}\n"
    )
    if val.suggestion:
        text += f"💡 建議：{val.suggestion}\n"
    text += "\n請選擇處理方式："
    await _tg_send(run.telegram_chat_id, text, _decision_keyboard(run.run_id))


async def _notify_final(run: PipelineRun, config: PipelineConfig):
    """發送 pipeline 最終結果摘要"""
    total = len(config.steps)
    ok_count = sum(1 for r in run.step_results if r.validation_status == "ok")

    status_map = {
        "completed": ("✅", "Pipeline 完成"),
        "aborted":   ("🛑", "Pipeline 已中止"),
    }
    emoji, title = status_map.get(run.status, ("❌", "Pipeline 失敗"))

    duration = ""
    if run.ended_at and run.started_at:
        try:
            secs = int((
                datetime.fromisoformat(run.ended_at) -
                datetime.fromisoformat(run.started_at)
            ).total_seconds())
            duration = f"⏱ 耗時：{secs // 60}m {secs % 60}s\n"
        except Exception:
            pass

    # Step 摘要
    step_lines = []
    for i, step in enumerate(config.steps):
        if i < len(run.step_results):
            r = run.step_results[i]
            icon = {"ok": "✅", "warning": "⚠️", "failed": "❌"}.get(r.validation_status, "❓")
            step_lines.append(f"  {icon} {step.name}")
        else:
            step_lines.append(f"  ⬜ {step.name}（未執行）")

    text = (
        f"{emoji} <b>{title}</b>\n\n"
        f"📋 {run.pipeline_name}\n"
        f"🔢 {ok_count}/{total} 步驟成功\n"
        f"{duration}"
        f"\n<b>步驟概覽：</b>\n" + "\n".join(step_lines) +
        f"\n\n📁 <code>{run.log_path}</code>"
    )
    await _tg_send(run.telegram_chat_id, text)


# ── Deterministic validation (fast recipe mode) ──────────────────────────────

def _deterministic_validate(step, exec_result, logger) -> ValidationResult:
    """Recipe 快速模式：不叫 LLM，只做確定性檢查。"""
    from pathlib import Path as _Path

    # 1. exit code
    if exec_result.exit_code != 0:
        return ValidationResult(
            status="failed",
            reason=f"Exit code {exec_result.exit_code}",
            suggestion="Recipe 執行失敗，建議改用完整模式重跑",
        )

    # 2. 輸出檔存在 + 大小
    if step.output and step.output.path:
        p = _Path(step.output.path).expanduser()
        if not p.exists():
            return ValidationResult(
                status="failed",
                reason=f"輸出檔案 {step.output.path} 不存在",
                suggestion="Recipe 未產生預期檔案，建議改用完整模式",
            )
        size = p.stat().st_size
        if size == 0:
            return ValidationResult(
                status="failed",
                reason=f"輸出檔案 {step.output.path} 為空檔案（0 bytes）",
                suggestion="Recipe 產生了空檔案，建議改用完整模式",
            )
        # CSV: 檢查有 header
        if p.suffix.lower() == ".csv":
            try:
                with open(p, "r", encoding="utf-8") as f:
                    lines = sum(1 for _ in f)
                if lines < 2:
                    return ValidationResult(
                        status="failed",
                        reason=f"CSV 檔案只有 {lines} 行（預期至少有 header + 資料）",
                        suggestion="",
                    )
            except Exception:
                pass
        # Excel: 檢查有 sheet
        if p.suffix.lower() in (".xlsx", ".xls"):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(p, read_only=True)
                sheet_count = len(wb.sheetnames)
                wb.close()
                if sheet_count == 0:
                    return ValidationResult(
                        status="failed",
                        reason="Excel 檔案沒有任何工作表",
                        suggestion="",
                    )
            except Exception:
                pass

    logger.info(f"[{step.name}] ⚡ 確定性檢查通過")
    return ValidationResult(
        status="ok",
        reason="exit code=0、輸出檔案存在且非空",
        suggestion="",
    )


def get_run_log_tail(run_id: str, lines: int = 30) -> str:
    """取得 pipeline 執行 log 的最後 N 行（供 Telegram 查看）"""
    store = get_store()
    run = store.load(run_id)
    if not run or not run.log_path:
        return "（找不到 log）"
    from pathlib import Path as _Path
    log_file = _Path(run.log_path)
    if not log_file.exists():
        return "（log 檔案不存在）"
    try:
        all_lines = log_file.read_text(encoding="utf-8").splitlines()
        tail = all_lines[-lines:] if len(all_lines) > lines else all_lines
        return "\n".join(tail)
    except Exception as e:
        return f"（讀取失敗：{e}）"


# ── Main pipeline engine ──────────────────────────────────────────────────────

async def run_pipeline(
    config_dict: dict,
    chat_id: int,
    run_id: Optional[str] = None,
    start_from_step: int = 0,
) -> str:
    """
    執行（或恢復）一個 pipeline。
    """
    store = get_store()

    # 建立或恢復 run
    if run_id:
        run = store.load(run_id)
        if not run:
            raise ValueError(f"找不到 pipeline run: {run_id}")
        
        # 確保 run 物件同步使用傳入的最新配置（包含 hint）
        run.config_dict = config_dict
        run.status = "running"
        run.current_step = start_from_step
        # 附加到原始 log 檔（不建新檔），前端讀到的 log_path 保持不變
        logger = resume_run_logger(run.run_id, run.log_path)
        logger.info(f"恢復執行，從步驟 {start_from_step + 1} 繼續")
    else:
        # 新建 run
        config = PipelineConfig.from_dict(config_dict)
        run_id = str(uuid.uuid4())[:12]
        logger, log_path = create_run_logger(run_id, config.name)
        run = PipelineRun(
            run_id=run_id,
            pipeline_name=config.name,
            config_dict=config_dict,
            telegram_chat_id=chat_id,
            log_path=log_path,
        )
        logger.info(f"Pipeline 開始：{config.name}，共 {len(config.steps)} 步驟")

    config = PipelineConfig.from_dict(run.config_dict)
    use_recipe = run.config_dict.get("_use_recipe", False)
    workflow_id = run.config_dict.get("_workflow_id") or run.workflow_id
    store.save(run)

    # ── Step loop ────────────────────────────────────────────
    completed_outputs: list[dict] = []  # 收集前步驟的輸出資訊

    # 恢復執行時，重建已完成步驟的輸出資訊（供後續步驟參考）
    if start_from_step > 0:
        from pathlib import Path as _Path
        # 預先用 step_index 索引 step_results、給沒設 output.path 的步驟回填 actual_output_path
        _sr_idx = {sr.step_index: sr for sr in run.step_results}
        for i in range(start_from_step):
            prev_step = config.steps[i] if i < len(config.steps) else None
            if not prev_step or prev_step.human_confirm:
                continue
            # 優先：明確 output.path > 該步 StepResult.actual_output_path
            _eff_path = ""
            if prev_step.output and prev_step.output.path:
                _eff_path = prev_step.output.path
            else:
                _sr = _sr_idx.get(i)
                if _sr and getattr(_sr, "actual_output_path", ""):
                    _eff_path = _sr.actual_output_path
            if _eff_path:
                p = _Path(_eff_path).expanduser()
                out_info = {"path": str(p), "schema": ""}
                try:
                    if p.suffix == ".csv" and p.exists():
                        with open(p, "r") as f:
                            out_info["schema"] = f.readline().strip()
                    elif p.suffix in (".xlsx", ".xls") and p.exists():
                        out_info["schema"] = "Excel 工作簿"
                    elif p.suffix in (".png", ".jpg", ".jpeg") and p.exists():
                        out_info["schema"] = "圖片檔案"
                except Exception:
                    pass
                completed_outputs.append(out_info)
        if completed_outputs:
            logger.info(f"已重建 {len(completed_outputs)} 個前步驟的輸出資訊：{[o['path'] for o in completed_outputs]}")

    no_save_recipe = run.config_dict.get("_no_save_recipe", False)

    while run.current_step < len(config.steps):
        # ── 每步開始前檢查中止旗標 ──
        if is_abort_requested(run.run_id):
            clear_abort(run.run_id)
            unregister_task(run.run_id)
            run.status = "aborted"
            run.ended_at = datetime.now().isoformat()
            store.save(run)
            logger.info("用戶透過 UI 中止 Pipeline")
            await _notify_final(run, config)
            return run.run_id

        step = config.steps[run.current_step]
        step_num = run.current_step + 1
        total = len(config.steps)
        logger.info(f"══ 步驟 {step_num}/{total}：{step.name} ══")

        # 步驟開始前 snapshot workflow 輸出資料夾（mtime 比對用）
        # 步驟結束後 _diff_snapshot_pick_main 找新增/修改的檔，存進 StepResult.actual_output_path
        # → TG「取任一步輸出」就能對應到該步真正寫的檔（不再讓多個 skill 步驟搶到「最新檔」）
        _step_dir_snapshot_before = _snapshot_workflow_dir(config.name)

        # ── 人工確認節點：暫停等待確認 ──
        if step.human_confirm:
            logger.info(f"[{step.name}] ✋ 人工確認節點，暫停等待確認")

            # 收集前一步結果摘要
            prev_summary = ""
            if run.step_results:
                prev = run.step_results[-1]
                status_icon = {"ok": "✅", "failed": "❌"}.get(prev.validation_status, "⚠️")
                prev_summary = (
                    f"前一步驟：{prev.step_name}\n"
                    f"狀態：{status_icon} {prev.validation_status}\n"
                    f"原因：{prev.validation_reason or '（無）'}\n"
                )
                if prev.stdout_tail:
                    prev_summary += f"輸出摘要：{prev.stdout_tail[-300:]}\n"

            confirm_msg = step.message or "請確認上一步結果是否正確，再繼續執行"
            full_message = f"{prev_summary}\n📋 {confirm_msg}"

            run.status = "awaiting_human"
            run.awaiting_type = "human_confirm"
            run.awaiting_message = confirm_msg

            # step_result 跟 status 一起一次寫完，後面 await TG / preview 期間就不再 save。
            # 否則用戶按通過時 resume_pipeline 把狀態改成 running，本協程的 stale run 物件
            # 在後面 store.save 又把狀態蓋回 awaiting_human → 用戶第二次按通過會再過 gate
            # → 同一步驟被啟動兩次（race condition，造成工作流跑慢／重複執行）。
            step_result = StepResult(
                step_index=run.current_step,
                step_name=step.name,
                exit_code=0,
                stdout_tail="等待人工確認",
                stderr_tail="",
                validation_status="ok",
                validation_reason="人工確認節點 — 等待中",
                validation_suggestion="",
                retries_used=0,
            )
            if len(run.step_results) > run.current_step:
                run.step_results[run.current_step] = step_result
            else:
                run.step_results.append(step_result)
            store.save(run)

            # 判斷「補充指示」按鈕要不要給：只有上一個可執行節點是 skill_mode 才顯示
            # （往回跳過其他連續的 human_confirm 節點，找真正要被重做的 step）
            _prev = run.current_step - 1
            while _prev >= 0 and config.steps[_prev].human_confirm:
                _prev -= 1
            allow_hint = _prev >= 0 and bool(config.steps[_prev].skill_mode)

            # Telegram 通知
            if step.notify_telegram:
                tg_text = (
                    f"✋ <b>Pipeline 等待確認</b>\n\n"
                    f"📋 {run.pipeline_name}\n"
                    f"📍 步驟 {step_num}/{total}：<b>{step.name}</b>\n\n"
                )
                if prev_summary:
                    tg_text += f"{prev_summary}\n"
                tg_text += f"💬 {confirm_msg}\n\n請選擇："
                await _tg_send(run.telegram_chat_id, tg_text,
                               _confirm_keyboard(run.run_id, screenshot=step.screenshot,
                                                 allow_hint=allow_hint,
                                                 preview_enabled=step.preview_prev_output))
                # 自動傳上一步輸出檔：step.send_prev_output=True 時，立刻把上一步 output.path
                # 的檔案（或資料夾打包成 zip）當 document 傳到 TG。手機上可直接點開 / 下載。
                # 失敗（沒設輸出 / 檔案不存在 / 太大）也只 log warning、不擋人工確認流程。
                if step.send_prev_output:
                    try:
                        # 找上一個非 human_confirm 步驟（連續多個 human_confirm 時往前跳過）
                        _po_prev = run.current_step - 1
                        while _po_prev >= 0 and config.steps[_po_prev].human_confirm:
                            _po_prev -= 1
                        if _po_prev >= 0:
                            _po_step = config.steps[_po_prev]
                            _po_result = next((sr for sr in run.step_results if sr.step_index == _po_prev), None)
                            ok, msg = await _send_step_output_to_tg(
                                run.telegram_chat_id, _po_step,
                                step_label=f"步驟 {_po_prev+1}：{_po_step.name}",
                                workflow_name=config.name,
                                logger=logger,
                                step_result=_po_result,
                            )
                            if ok:
                                logger.info(f"[{step.name}] ✓ 自動傳上一步輸出：{msg}")
                            else:
                                # 不再廣播警告到 TG（noise）：按鈕「📎 上一步輸出」永遠存在，
                                # 使用者要時自己點即可。失敗只 log 到 backend、debug 用。
                                logger.warning(
                                    f"[{step.name}] 自動傳上一步輸出未成功（不廣播到 TG）：{msg}"
                                )
                    except Exception as _e:
                        logger.warning(f"[{step.name}] send_prev_output 例外：{_e}")

                # 自動截圖：step.screenshot=True 時，發完決策訊息立刻截全螢幕附過去，
                # 逐螢幕送（雙螢幕 → 2 張，方便 TG 上直接看到上一步結果不用再按按鈕）
                if step.screenshot:
                    try:
                        ss_paths = take_screenshots(run.pipeline_name, step.name)
                        if ss_paths:
                            await _tg_send_photos(
                                run.telegram_chat_id,
                                ss_paths,
                                caption_prefix=f"📸 {step.name}",
                            )
                    except Exception as _e:
                        logger.warning(f"[{step.name}] 自動截圖傳送失敗：{_e}")

                # 檔案預覽：preview_prev_output=True 時，把上一步的 output.path 檔案
                # render 成 PNG 一併傳，讓使用者在手機上直接看到內容（不用 SSH 回電腦）
                # B1 路線：pandas/python-docx/python-pptx/pypdfium2/PIL 純 headless；
                # 後備：LibreOffice 無頭轉 PDF（需使用者自己裝 soffice）
                if step.preview_prev_output:
                    prev_file = _find_prev_output_file(run, config)
                    if not prev_file:
                        logger.info(f"[{step.name}] preview_prev_output 開啟但找不到上一步輸出檔，跳過")
                    else:
                        try:
                            # render 同步跑（pandas/PIL 不算慢；LibreOffice 若用到才可能 >5s）
                            # 放 executor 裡別 block event loop
                            from pipeline.file_preview import render_file_preview
                            preview_paths = await asyncio.get_event_loop().run_in_executor(
                                None,
                                lambda fp=prev_file: render_file_preview(fp, out_dir=str(Path(fp).parent)),
                            )
                            if preview_paths:
                                logger.info(f"[{step.name}] 📄 預覽產生 {len(preview_paths)} 張 → 傳 TG")
                                await _tg_send_photos(
                                    run.telegram_chat_id,
                                    preview_paths,
                                    caption_prefix=f"📄 上一步驟輸出預覽：{Path(prev_file).name}",
                                )
                            else:
                                logger.warning(f"[{step.name}] 預覽 render 回傳空清單：{prev_file}")
                        except Exception as _e:
                            logger.warning(f"[{step.name}] 檔案預覽失敗：{_e}")

            # step_result 已於 await 前寫入；這裡只純粹釋放 task 並退出協程
            unregister_task(run.run_id)
            return run.run_id  # 暫停，等 resume_pipeline 被呼叫

        logger.debug(f"[{step.name}] batch 全文（{len(step.batch)} 字元）：{step.batch[:500]}")

        retries_used = 0
        step_failures: list[dict] = []  # 累積此步驟的失敗歷史，傳給下次重試

        # 計算當前步驟的工作目錄 (Working Directory)
        from pathlib import Path as _Path
        # 定義專案根目錄 (backend/pipeline/runner.py 的上三層)
        _PROJ_ROOT = _Path(__file__).parent.parent.parent.absolute()

        def _resolve_path(p: str) -> _Path:
            """把 output.path 解析成絕對路徑：
            - `~/xxx` 展開到使用者家目錄
            - 絕對路徑直接用
            - 相對路徑 → 以**專案根目錄**為基準（而非 backend cwd）"""
            pp = _Path(p).expanduser()
            if not pp.is_absolute():
                pp = _PROJ_ROOT / pp
            return pp

        # 預設：專案根目錄/ai_output/{pipeline_name}/
        default_wd = str(_PROJ_ROOT / "ai_output" / config.name)
        wd = step.working_dir
        if not wd and step.output and step.output.path:
            wd = str(_resolve_path(step.output.path).parent)
        if not wd:
            wd = default_wd
        _Path(wd).mkdir(parents=True, exist_ok=True)

        # Retry loop for this step
        while True:
            if step.visual_validation:
                # ── 視覺驗證節點：純 VLM 判斷，不執行命令 ──
                from .visual_validator import run_visual_validation
                from .executor import ExecResult as _ExecResult
                prev_file = _find_prev_output_file(run, config) if step.vv_source != "current_screen" else None
                # search_region 解析：4 整數 [l,t,w,h]，否則 None
                _vsr = step.vv_search_region or []
                vv_region = None
                if isinstance(_vsr, (list, tuple)) and len(_vsr) == 4:
                    try:
                        vv_region = (int(_vsr[0]), int(_vsr[1]), int(_vsr[2]), int(_vsr[3]))
                        if vv_region[2] <= 0 or vv_region[3] <= 0:
                            vv_region = None
                    except (TypeError, ValueError):
                        vv_region = None
                vv_pass, vv_reason = await run_visual_validation(
                    source=step.vv_source,
                    prompt=step.vv_prompt,
                    prev_output_file=prev_file,
                    out_dir=wd,
                    search_region=vv_region,
                    logger=logger,
                )
                exec_result = _ExecResult(
                    exit_code=0 if vv_pass else 1,
                    stdout=f"[visual_validation] {vv_reason}",
                    stderr="" if vv_pass else f"VLM 判斷未通過：{vv_reason}",
                )
            elif step.computer_use:
                # ── 桌面自動化節點：純 pyautogui + opencv，不走 LLM / recipe ──
                from .computer_use import execute_computer_use_step
                from .executor import ExecResult as _ExecResult
                # assets_dir 相對路徑解析：若為空，預設 ai_output/<pipeline>/<step_name>_assets
                if step.assets_dir:
                    assets_abs = str(_resolve_path(step.assets_dir))
                else:
                    assets_abs = str(_resolve_path(f"ai_output/{config.name}/{step.name}_assets"))
                # actions 是 ComputerUseAction pydantic model，轉成 dict list 傳進引擎
                # by_alias=True 確保 else_ 這類為了閃 Python 保留字取的別名，
                # 在轉 dict 時還原為 YAML 原生的 "else" key（讓 execute_action 用 .get("else") 讀得到）
                actions_dicts = [a.model_dump(by_alias=True) if hasattr(a, "model_dump") else dict(a) for a in (step.actions or [])]
                _cu_result = await asyncio.get_event_loop().run_in_executor(
                    None,
                    lambda: execute_computer_use_step(
                        actions=actions_dicts,
                        assets_dir=assets_abs,
                        logger=logger,
                        run_id=run.run_id,
                        fail_fast=step.fail_fast,
                        cv_threshold=step.cv_threshold,
                        cv_search_only_near=step.cv_search_only_near,
                        cv_search_radius=step.cv_search_radius,
                        cv_trigger_hover=step.cv_trigger_hover,
                        cv_hover_wait_ms=step.cv_hover_wait_ms,
                        cv_coord_fallback=step.cv_coord_fallback,
                        ocr_threshold=step.ocr_threshold,
                        ocr_cv_fallback=step.ocr_cv_fallback,
                    ),
                )
                # 映射回 ExecResult 讓後續驗證/重試邏輯通用
                exec_result = _ExecResult(
                    exit_code=_cu_result.exit_code,
                    stdout=_cu_result.stdout,
                    stderr=_cu_result.stderr,
                )
            elif step.skill_mode:
                # recipe key 使用「索引:名稱」避免同名步驟互相覆蓋
                recipe_step_key = f"{step_num}:{step.name}"
                # 把 output_path 解析成絕對路徑傳給 LLM，避免 LLM 搞不清楚相對於哪個 cwd
                _resolved_out = str(_resolve_path(step.output.path)) if (step.output and step.output.path) else None
                exec_result = await execute_step_with_skill(
                    task_description=step.batch,
                    timeout=step.timeout,
                    logger=logger,
                    step_name=step.name,
                    output_path=_resolved_out,
                    working_dir=wd,
                    prev_outputs=completed_outputs if completed_outputs else None,
                    pipeline_id=workflow_id or config.name,
                    use_recipe=use_recipe,
                    no_save_recipe=no_save_recipe,
                    readonly=step.readonly,
                    run_id=run.run_id,
                    previous_failures=step_failures if step_failures else None,
                    recipe_step_key=recipe_step_key,
                    skill_name=step.skill,
                    ask_mode=step.ask_mode,
                )
            else:
                exec_result = await execute_step(
                    command=step.batch,
                    timeout=step.timeout,
                    logger=logger,
                    step_name=step.name,
                    run_id=run.run_id,
                    working_dir=wd,
                )

            # 快速模式：Recipe 命中 + 執行成功 → 確定性驗證（不叫 LLM）
            recipe_hit = (exec_result.stderr == "__RECIPE_HIT__")
            if recipe_hit:
                exec_result.stderr = ""  # 清掉標記

            has_expect = step.output and step.output.get_expect()
            # exit_code -429 = LLM 配額用盡（executor 標記），直接走 rate_limited 路徑、不再叫 validator（會再 429 一次）
            if exec_result.exit_code == -429:
                val = ValidationResult(
                    status="rate_limited",
                    reason=(exec_result.stderr or "LLM 配額用盡或速率受限（429）"),
                    suggestion="等配額重置或在 Settings 切換 provider（Groq / OpenRouter / Ollama 本地）",
                )
            # visual_validation 節點：節點自己就是 VLM 判斷，不需要再跑一次 LLM 驗證
            elif step.visual_validation:
                _status = "ok" if exec_result.exit_code == 0 else "failed"
                val = ValidationResult(
                    status=_status,
                    reason=exec_result.stdout.replace("[visual_validation] ", "") or "視覺驗證",
                    suggestion=exec_result.stderr if _status == "failed" else "",
                )
            # computer_use 節點：成敗已由 action 執行結果決定，不需 LLM 驗證
            elif step.computer_use:
                _status = "ok" if exec_result.exit_code == 0 else "failed"
                val = ValidationResult(
                    status=_status,
                    reason=f"桌面自動化 {exec_result.stdout.count('OK')} 個動作成功"
                           + (f"，{exec_result.exit_code} 個失敗" if exec_result.exit_code != 0 else ""),
                    suggestion=exec_result.stderr if _status == "failed" else "",
                )
            elif recipe_hit and use_recipe and exec_result.exit_code == 0 and not has_expect:
                # 確定性檢查：exit code=0、輸出檔存在、檔案大小合理（無 AI 驗證節點）
                val = _deterministic_validate(step, exec_result, logger)
            elif recipe_hit and use_recipe and exec_result.exit_code == 0 and has_expect:
                # Recipe 命中但有 AI 驗證節點 → 快速 LLM 驗證（不走 Skill 深度驗證）
                logger.info(f"[{step.name}] 🔍 Recipe 命中 + 有 AI 驗證需求，走快速 LLM 驗證")
                val = await validate_step(
                    step_name=step.name,
                    command=step.batch,
                    exit_code=exec_result.exit_code,
                    stdout=exec_result.stdout,
                    stderr=exec_result.stderr,
                    output_path=(str(_resolve_path(step.output.path)) if (step.output and step.output.path) else None),
                    output_expect=step.output.get_expect() if step.output else None,
                    logger=logger,
                )
            elif config.validate and has_expect:
                # 使用者填了「預期輸出描述」→ 跑 LLM 驗證
                # output.skill_mode=true 走深度（agent 主動跑工具驗證）；否則走快速 LLM 一次驗證
                use_skill = step.output and step.output.skill_mode
                validate_fn = validate_step_with_skill if use_skill else validate_step
                val = await validate_fn(
                    step_name=step.name,
                    command=step.batch,
                    exit_code=exec_result.exit_code,
                    stdout=exec_result.stdout,
                    stderr=exec_result.stderr,
                    output_path=(str(_resolve_path(step.output.path)) if (step.output and step.output.path) else None),
                    output_expect=step.output.get_expect() if step.output else None,
                    logger=logger,
                )
            elif config.validate and not has_expect and step.skill_mode:
                # Skill 節點沒填 expect → 仍走 LLM 淺驗證
                # 理由：LLM 寫的程式碼容易 silent fail（exit_code=0 但結果語意錯，例如
                # 「抓 10 篇」實際只抓到 3 篇），確定性檢查抓不到、需要外層 LLM 看內容把關。
                # 跑 validate_step（淺、單次 LLM call ~5-15s），不走 skill 深度模式。
                logger.info(f"[{step.name}] 🔍 Skill 節點預設驗證（沒填 expect、走淺 LLM 把關 silent fail）")
                val = await validate_step(
                    step_name=step.name,
                    command=step.batch,
                    exit_code=exec_result.exit_code,
                    stdout=exec_result.stdout,
                    stderr=exec_result.stderr,
                    output_path=(str(_resolve_path(step.output.path)) if (step.output and step.output.path) else None),
                    output_expect=None,
                    logger=logger,
                )
            elif config.validate and not has_expect:
                # Script / 其他無 skill_mode 節點 + 沒填 expect → 只做確定性檢查
                # 理由：script 是使用者自己寫的程式、自己負責正確性，外層 LLM 沒足夠上下文判斷
                val = _deterministic_validate(step, exec_result, logger)
                logger.info(f"[{step.name}] ⚡ Script 節點沒填預期輸出，只看 exit code + 檔案存在")
            else:
                status = "ok" if exec_result.exit_code == 0 else "failed"
                val = ValidationResult(
                    status=status,
                    reason=f"Exit code {exec_result.exit_code}（LLM 驗證已停用）",
                    suggestion="" if status == "ok" else "請查看 log 取得詳細錯誤",
                )
                logger.info(f"[{step.name}] 驗證（僅 exit code）：{val.status}")

            # ── 算這步真正寫到 workflow dir 的主要檔案 ─────────────────
            # 優先順序：明確 output.path > snapshot diff > 空字串
            # snapshot diff 對沒設 output.path 的 skill 步驟最關鍵 —
            # 否則 TG「取任一步輸出」會多步搶同一個「workflow dir 最新檔」
            actual_out = ""
            try:
                if step.output and step.output.path:
                    p = _resolve_path(step.output.path)
                    if p.exists() and p.is_file():
                        actual_out = str(p.absolute())
                if not actual_out:
                    actual_out = _diff_snapshot_pick_main(_step_dir_snapshot_before, config.name) or ""
            except Exception as _e:
                logger.debug(f"[{step.name}] snapshot diff 失敗（略過）：{_e}")

            step_result = StepResult(
                step_index=run.current_step,
                step_name=step.name,
                exit_code=exec_result.exit_code,
                stdout_tail=exec_result.stdout[-500:],
                stderr_tail=exec_result.stderr[-200:],
                validation_status=val.status,
                validation_reason=val.reason,
                validation_suggestion=val.suggestion,
                retries_used=retries_used,
                actual_output_path=actual_out,
            )

            # 更新或追加步驟結果
            if len(run.step_results) > run.current_step:
                run.step_results[run.current_step] = step_result
            else:
                run.step_results.append(step_result)
            store.save(run)

            # rate_limited（LLM provider 429）：立即暫停，不重試（重試只會再 429 燒配額）
            # 跳到 awaiting_human，讓使用者決定等多久 / 切 provider / 中止
            if val.status == "rate_limited":
                logger.warning(f"步驟 {step_num} ⏸ LLM 配額用盡（429）— 暫停等使用者決策，不重試")
                run.status = "awaiting_human"
                run.awaiting_type = "rate_limited"
                run.awaiting_message = (
                    f"⚠ LLM provider 配額用盡或速率受限（429）\n\n"
                    f"原因：{val.reason}\n\n"
                    f"建議：{val.suggestion}\n\n"
                    f"請選擇：等待重試 / 切換 provider 後重試 / 中止"
                )
                store.save(run)
                return run.run_id

            if val.status == "ok":
                logger.info(f"步驟 {step_num} ✅ 通過")
                # 收集延遲儲存的 recipe
                if hasattr(exec_result, 'pending_recipe') and exec_result.pending_recipe:
                    run.pending_recipes.append(exec_result.pending_recipe)
                # 收集此步驟的輸出資訊供後續步驟參考
                # 優先：明確 output.path > snapshot 算出來的 actual_output_path
                # 後者讓沒設 output.path 的 skill 步驟也能被後續步驟自動抓到正確檔
                _eff_path = ""
                if step.output and step.output.path:
                    _eff_path = step.output.path
                elif actual_out:  # 上面 snapshot diff 算出來的
                    _eff_path = actual_out
                if _eff_path:
                    out_info = {"path": _eff_path, "schema": ""}
                    try:
                        from pathlib import Path as _Path
                        p = _Path(_eff_path)
                        if p.suffix == ".csv" and p.exists():
                            with open(p, "r") as f:
                                header = f.readline().strip()
                            out_info["schema"] = header
                        elif p.suffix in (".xlsx", ".xls") and p.exists():
                            out_info["schema"] = "Excel 工作簿"
                        elif p.suffix in (".png", ".jpg", ".jpeg") and p.exists():
                            out_info["schema"] = "圖片檔案"
                    except Exception:
                        pass
                    completed_outputs.append(out_info)
                run.current_step += 1
                store.save(run)
                break  # 進入下一步

            elif retries_used < step.retry:
                retries_used += 1
                # 記錄此次失敗的原因與建議，供下次重試時傳給 LLM
                step_failures.append({
                    "attempt": retries_used,
                    "reason": val.reason,
                    "suggestion": val.suggestion,
                    "stdout_tail": exec_result.stdout[-800:] if exec_result.stdout else "",
                    "stderr_tail": exec_result.stderr[-400:] if exec_result.stderr else "",
                })
                logger.warning(
                    f"步驟 {step_num} 驗證失敗，自動重試 {retries_used}/{step.retry}：{val.reason}"
                )
                continue  # 重試

            else:
                # 重試耗盡，暫停等待人為決策
                logger.warning(f"步驟 {step_num} 失敗且重試次數耗盡，等待人為決策")
                run.status = "awaiting_human"
                run.awaiting_type = "failure"
                run.awaiting_message = val.reason or ""

                # 優先使用 LLM 回報的 missing_packages 建立具體安裝建議
                missing_pkgs = getattr(exec_result, 'missing_packages', None) or []
                # 也嘗試從 stderr 偵測 ModuleNotFoundError
                if not missing_pkgs and exec_result.stderr:
                    import re as _re
                    found = _re.findall(r"ModuleNotFoundError: No module named '([^']+)'", exec_result.stderr)
                    if found:
                        missing_pkgs = list(dict.fromkeys(found))  # 去重保序

                if missing_pkgs:
                    pkgs_str = "、".join(missing_pkgs)
                    install_hint = f"建議在「設定 → 套件管理」安裝以下套件後再重試：{pkgs_str}"
                    run.awaiting_suggestion = install_hint + (f"\n\nAI 說明：{val.suggestion}" if val.suggestion else "")
                else:
                    run.awaiting_suggestion = val.suggestion or ""

                store.save(run)
                await _notify_failure(run, val, step.name)
                unregister_task(run.run_id)
                return run.run_id  # 暫停

    # ── 全部步驟完成 ─────────────────────────────────────────
    clear_abort(run.run_id)
    unregister_task(run.run_id)
    run.status = "completed"
    run.ended_at = datetime.now().isoformat()
    store.save(run)
    logger.info(f"Pipeline {config.name} 全部完成！")
    await _notify_final(run, config)
    return run.run_id


# ── Human-in-the-loop resume ─────────────────────────────────────────────────

async def resume_pipeline(run_id: str, decision: str, hint: str = "") -> str:
    """
    用戶透過 Telegram inline keyboard 做出決策後，呼叫此函式繼續執行。

    Args:
        run_id:   pipeline run id
        decision: "retry" | "skip" | "abort" | "continue" | "retry_with_hint"
        hint:     補充指示（retry_with_hint 時使用）

    Returns:
        str 回應訊息（回覆給用戶）
    """
    store = get_store()
    run = store.load(run_id)

    if not run:
        return f"❌ 找不到 Pipeline run：{run_id}"
    if run.status != "awaiting_human":
        return f"⚠️ Pipeline {run_id} 目前狀態為 {run.status}，無需決策"

    config = PipelineConfig.from_dict(run.config_dict)
    step_num = run.current_step + 1
    total = len(config.steps)
    # 附加到原始 log 檔，確保前端讀到的 log_path 始終指向同一個檔案
    logger = resume_run_logger(run.run_id, run.log_path)

    # ── ask_user 回答：skill agent 仍在 in-memory 等待 event ──
    if run.awaiting_type == "ask_user":
        from pipeline.executor import deliver_ask_user_answer
        if decision == "answer":
            ok = deliver_ask_user_answer(run_id, hint)
            if not ok:
                # agent 可能已 timeout 或後端已重啟
                return "⚠️ 答案送達失敗：skill agent 已不在等待狀態（可能逾時或後端重啟）"
            logger.info(f"[ask_user] 使用者答案已送達：{hint[:100]}")
            return f"✅ 答案已送出"
        elif decision == "abort":
            # 先中止 skill agent 的等待（讓它收到 None），再把 pipeline 標為 aborted
            deliver_ask_user_answer(run_id, "")  # 空字串讓 agent 繼續但不拿到答案
            run.status = "aborted"
            run.ended_at = datetime.now().isoformat()
            store.save(run)
            logger.info("[ask_user] 使用者選擇中止")
            await _notify_final(run, config)
            return f"🛑 Pipeline 已中止"
        else:
            return f"⚠️ ask_user 只接受 answer 或 abort，收到 {decision}"

    if decision == "abort":
        run.status = "aborted"
        run.ended_at = datetime.now().isoformat()
        store.save(run)
        logger.info("用戶選擇中止 Pipeline")
        await _notify_final(run, config)
        return f"🛑 Pipeline 已中止（步驟 {step_num}/{total}）"

    elif decision == "skip":
        logger.info(f"用戶選擇跳過步驟 {step_num}")
        next_step = run.current_step + 1

        if next_step >= total:
            run.status = "completed"
            run.ended_at = datetime.now().isoformat()
            store.save(run)
            await _notify_final(run, config)
            return f"⏩ 跳過最後一步，Pipeline 完成"

        run.awaiting_type = ""
        run.awaiting_message = ""
        run.awaiting_suggestion = ""
        run.status = "running"
        store.save(run)

        async def _delayed_skip():
            await asyncio.sleep(0.2)
            t = asyncio.create_task(run_pipeline(
                config_dict=run.config_dict,
                chat_id=run.telegram_chat_id,
                run_id=run.run_id,
                start_from_step=next_step,
            ))
            register_task(run.run_id, t)

        asyncio.create_task(_delayed_skip())
        return f"⏩ 跳過步驟 {step_num}，繼續執行步驟 {step_num + 1}/{total}"

    elif decision == "retry":
        logger.info(f"用戶選擇重試步驟 {step_num}")
        run.awaiting_type = ""
        run.awaiting_message = ""
        run.awaiting_suggestion = ""
        run.status = "running"
        store.save(run)

        async def _delayed_retry():
            await asyncio.sleep(0.2)
            t = asyncio.create_task(run_pipeline(
                config_dict=run.config_dict,
                chat_id=run.telegram_chat_id,
                run_id=run.run_id,
                start_from_step=run.current_step,
            ))
            register_task(run.run_id, t)

        asyncio.create_task(_delayed_retry())
        return f"🔄 重試步驟 {step_num}/{total}"

    elif decision == "retry_with_hint":
        import copy
        # 1. 使用深拷貝，確保 config 修改是獨立且完整的
        config_d = copy.deepcopy(run.config_dict)
        steps = config_d.get("steps", [])

        is_confirm = run.awaiting_type == "human_confirm"
        target = run.current_step

        if is_confirm:
            prev_step = run.current_step - 1
            while prev_step >= 0 and steps[prev_step].get("human_confirm"):
                prev_step -= 1
            if prev_step < 0:
                return "⚠️ 確認節點前沒有可重做的步驟"
            # 防呆：只有 skill_mode 節點能消化 hint。shell / computer_use 重跑 hint 無意義或會壞掉，
            # 正常 UI 不會給這個按鈕，但舊訊息或外部 API 呼叫還是可能打進來 → 拒絕
            if not steps[prev_step].get("skill_mode"):
                return (
                    "⚠️ 上一步不是 AI 技能節點，無法使用補充指示。"
                    "補充指示會附加給 LLM 重新生成程式碼；shell / 桌面自動化節點沒有 LLM 可消化。"
                )
            target = prev_step

        if target < len(steps):
            original_batch = steps[target].get("batch", "")
            # 清理舊的提示詞標籤，避免重複疊加
            clean_batch = original_batch.split("【用戶補充指示】")[0].strip()
            steps[target]["batch"] = f"{clean_batch}\n\n【用戶補充指示】{hint}"
            config_d["steps"] = steps

        # 2. 更新 run 狀態並「立即」同步回資料庫
        run.config_dict = config_d
        run.awaiting_type = ""
        run.awaiting_message = ""
        run.awaiting_suggestion = ""
        run.status = "running"
        store.save(run)

        # 3. 關鍵修正：給 Windows 一點點時間釋放資料庫鎖定
        async def _delayed_start():
            await asyncio.sleep(0.2)
            t = asyncio.create_task(run_pipeline(
                config_dict=config_d,  # 傳入已修改的配置
                chat_id=run.telegram_chat_id,
                run_id=run.run_id,
                start_from_step=target,
            ))
            register_task(run.run_id, t)

        asyncio.create_task(_delayed_start())
        
        if is_confirm:
            return f"💬 已附加指示，重做步驟 {target + 1}/{total}"
        else:
            return f"💬 已附加指示，重試步驟 {step_num}/{total}"

    elif decision == "continue":
        # 人工確認通過 → 繼續下一步
        logger.info(f"用戶確認通過步驟 {step_num}，繼續執行")

        # 更新確認步驟結果
        if run.current_step < len(run.step_results):
            run.step_results[run.current_step].validation_reason = "人工確認 — 已通過"
            run.step_results[run.current_step].stdout_tail = "已確認通過"

        run.awaiting_type = ""
        run.awaiting_message = ""
        next_step = run.current_step + 1

        if next_step >= total:
            run.status = "completed"
            run.ended_at = datetime.now().isoformat()
            store.save(run)
            logger.info(f"Pipeline {run.pipeline_name} 全部完成！")
            await _notify_final(run, config)
            return f"✅ 確認通過，Pipeline 全部完成"

        run.status = "running"
        store.save(run)

        async def _delayed_continue():
            await asyncio.sleep(0.2)
            t = asyncio.create_task(run_pipeline(
                config_dict=run.config_dict,
                chat_id=run.telegram_chat_id,
                run_id=run.run_id,
                start_from_step=next_step,
            ))
            register_task(run.run_id, t)

        asyncio.create_task(_delayed_continue())
        return f"✅ 確認通過，繼續執行步驟 {next_step + 1}/{total}"

    return "❓ 未知決策"
