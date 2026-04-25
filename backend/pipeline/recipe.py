"""Recipe Book — 將每次 Skill 模式成功產出的 run_python 快取起來，
下次若 task 描述 + 輸入檔案指紋都吻合，就直接跑 code 跳過 LLM。

資料位置：~/ai_output/pipeline_recipes/<pipeline_id_hash>/<step_name>.json
"""
from __future__ import annotations
import hashlib
import json
import logging
import threading
import time
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import Optional

from config import OUTPUT_BASE_PATH

logger = logging.getLogger(__name__)

_RECIPE_ROOT = OUTPUT_BASE_PATH / "pipeline_recipes"
_lock = threading.Lock()


def _sha1(text: str) -> str:
    return hashlib.sha1(text.encode("utf-8")).hexdigest()[:16]


def _safe_name(name: str) -> str:
    """Filesystem-safe 的 step 名稱。"""
    return "".join(c if c.isalnum() or c in "-_." else "_" for c in name)[:80]


def _fingerprint_input(path: str) -> str:
    """計算輸入檔案的輕量指紋（用 schema 而非內容 hash，才能容忍資料更新）。"""
    try:
        p = Path(path)
        if not p.exists():
            return f"missing:{path}"
        suffix = p.suffix.lower()
        if suffix == ".csv":
            with open(p, "r", encoding="utf-8", errors="replace") as f:
                header = f.readline().strip()
            return f"csv:{_sha1(header)}"
        if suffix in (".xlsx", ".xls"):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(p, read_only=True, data_only=True)
                sheets = "|".join(wb.sheetnames)
                wb.close()
                return f"xlsx:{_sha1(sheets)}"
            except Exception:
                # fallback: 固定標記（不用 size，因為資料量變化會導致 size 不穩定）
                return "xlsx:opaque"
        if suffix == ".json":
            try:
                with open(p, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    keys = "|".join(sorted(data.keys()))
                    return f"json:obj:{_sha1(keys)}"
                if isinstance(data, list) and data and isinstance(data[0], dict):
                    keys = "|".join(sorted(data[0].keys()))
                    return f"json:list:{_sha1(keys)}"
                return f"json:{type(data).__name__}"
            except Exception:
                return "json:parse_err"
        if suffix in (".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"):
            return f"image:{suffix[1:]}"
        # fallback：只用副檔名（不用 size，因為資料量變化會導致不穩定）
        return f"{suffix[1:] or 'bin'}:exists"
    except Exception as e:
        return f"error:{type(e).__name__}"


@dataclass
class Recipe:
    recipe_id: str
    pipeline_id: str
    step_name: str
    task_hash: str
    input_fingerprints: dict           # {path: fingerprint}
    output_path: Optional[str]
    code: str
    python_version: str = ""
    success_count: int = 0
    fail_count: int = 0
    created_at: float = field(default_factory=time.time)
    last_success_at: float = 0.0
    last_fail_at: float = 0.0
    avg_runtime_sec: float = 0.0
    disabled: bool = False

    def to_dict(self) -> dict:
        return asdict(self)

    @classmethod
    def from_dict(cls, d: dict) -> "Recipe":
        return cls(**{k: d.get(k) for k in cls.__dataclass_fields__.keys()})


def _recipe_path(pipeline_id: str, step_name: str) -> Path:
    pid_hash = _sha1(pipeline_id)
    return _RECIPE_ROOT / pid_hash / f"{_safe_name(step_name)}.json"


def load_recipe(pipeline_id: str, step_name: str) -> Optional[Recipe]:
    path = _recipe_path(pipeline_id, step_name)
    if not path.exists():
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return Recipe.from_dict(data)
    except Exception as e:
        logger.warning(f"Recipe 讀取失敗 {path}: {e}")
        return None


def match_recipe(
    pipeline_id: str,
    step_name: str,
    task_description: str,
    input_paths: list[str],
) -> Optional[Recipe]:
    """檢查是否有可重用的 recipe：task_hash 與輸入指紋都吻合，且沒被 disabled。"""
    r = load_recipe(pipeline_id, step_name)
    if r is None or r.disabled:
        return None
    if r.task_hash != _sha1(task_description):
        logger.info(f"[recipe] {step_name}：task_hash 不一致，需要重新學習")
        return None
    current_fp = {p: _fingerprint_input(p) for p in input_paths}
    if current_fp != r.input_fingerprints:
        logger.info(f"[recipe] {step_name}：輸入 schema 變了，需要重新學習")
        logger.debug(f"[recipe] 舊：{r.input_fingerprints}")
        logger.debug(f"[recipe] 新：{current_fp}")
        return None
    return r


def save_recipe(
    pipeline_id: str,
    step_name: str,
    task_description: str,
    input_paths: list[str],
    code: str,
    output_path: Optional[str],
    runtime_sec: float,
) -> Recipe:
    """成功執行後呼叫，寫入或更新 recipe。"""
    import sys
    with _lock:
        existing = load_recipe(pipeline_id, step_name)
        fps = {p: _fingerprint_input(p) for p in input_paths}
        task_hash = _sha1(task_description)

        if existing and existing.task_hash == task_hash and existing.input_fingerprints == fps and not existing.disabled:
            # 累積統計
            existing.success_count += 1
            existing.last_success_at = time.time()
            existing.code = code  # 更新成最新可運行的 code
            existing.avg_runtime_sec = (
                (existing.avg_runtime_sec * (existing.success_count - 1) + runtime_sec)
                / existing.success_count
            )
            r = existing
        else:
            r = Recipe(
                recipe_id=_sha1(f"{pipeline_id}:{step_name}:{task_hash}"),
                pipeline_id=pipeline_id,
                step_name=step_name,
                task_hash=task_hash,
                input_fingerprints=fps,
                output_path=output_path,
                code=code,
                python_version=f"{sys.version_info.major}.{sys.version_info.minor}",
                success_count=1,
                last_success_at=time.time(),
                avg_runtime_sec=runtime_sec,
            )
        path = _recipe_path(pipeline_id, step_name)
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(r.to_dict(), f, ensure_ascii=False, indent=2)
        logger.info(f"[recipe] 已儲存 {step_name} (success={r.success_count}, runtime={r.avg_runtime_sec:.1f}s)")
        return r


def mark_recipe_failed(pipeline_id: str, step_name: str) -> None:
    """Recipe 重跑失敗時標記，達到 3 次連續失敗就 disabled。"""
    with _lock:
        r = load_recipe(pipeline_id, step_name)
        if r is None:
            return
        r.fail_count += 1
        r.last_fail_at = time.time()
        if r.fail_count >= 3:
            r.disabled = True
            logger.warning(f"[recipe] {step_name}：連續失敗 {r.fail_count} 次，已停用")
        path = _recipe_path(pipeline_id, step_name)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(r.to_dict(), f, ensure_ascii=False, indent=2)


def list_recipes() -> list[dict]:
    """列出所有 recipe（用於前端管理頁）。"""
    if not _RECIPE_ROOT.exists():
        return []
    results = []
    for sub in _RECIPE_ROOT.iterdir():
        if not sub.is_dir():
            continue
        for f in sub.glob("*.json"):
            try:
                with open(f, "r", encoding="utf-8") as fh:
                    results.append(json.load(fh))
            except Exception:
                pass
    return results


def delete_recipe(pipeline_id: str, step_name: str) -> bool:
    path = _recipe_path(pipeline_id, step_name)
    if path.exists():
        path.unlink()
        return True
    return False


def get_pipeline_recipe_status(pipeline_id: str, step_names: list[str]) -> dict:
    """檢查 pipeline 的 recipe 覆蓋狀態。

    Returns:
        {
            "has_recipes": bool,        # 是否有任何 recipe
            "total_skill_steps": int,   # skill 模式的步驟數
            "covered_steps": int,       # 已有 recipe 的步驟數
            "steps": {step_name: {"has_recipe": bool, "success_count": int, ...}, ...}
        }
    """
    steps_info = {}
    covered = 0
    for name in step_names:
        r = load_recipe(pipeline_id, name)
        if r and not r.disabled:
            steps_info[name] = {
                "has_recipe": True,
                "success_count": r.success_count,
                "avg_runtime_sec": round(r.avg_runtime_sec, 1),
            }
            covered += 1
        else:
            steps_info[name] = {"has_recipe": False, "success_count": 0, "avg_runtime_sec": 0}
    return {
        "has_recipes": covered > 0,
        "total_skill_steps": len(step_names),
        "covered_steps": covered,
        "steps": steps_info,
    }


def delete_pipeline_recipes(pipeline_id: str) -> int:
    """刪除整個 pipeline 的所有 recipes。"""
    pid_hash = _sha1(pipeline_id)
    dir_path = _RECIPE_ROOT / pid_hash
    if not dir_path.exists():
        return 0
    count = 0
    for f in dir_path.glob("*.json"):
        f.unlink()
        count += 1
    try:
        dir_path.rmdir()
    except OSError:
        pass
    return count
