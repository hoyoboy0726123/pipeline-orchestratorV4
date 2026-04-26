"""
LLM 語意驗證器。

不靠關鍵字比對，讓 LLM 理解整體 log 內容，
判斷步驟是否真正成功——能區分「Python WARNING 不代表失敗」
與「真正的 Exception / 資料異常」。

支援：
- 文字檔讀取前 N 行供 LLM 判斷
- CSV / JSON / Excel 結構化摘要（欄位、行數、樣本）
- 圖片檔以 base64 傳給 Vision model 做視覺驗證
- Skill 模式：LLM 主動執行 Python / Shell 驗證程式碼（ReAct agent）
"""
import asyncio
import base64
import csv
import io
import json
import logging
import re
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from langchain_groq import ChatGroq
from langchain_core.messages import HumanMessage, SystemMessage

from config import GROQ_API_KEY, GROQ_MODEL_MAIN

# Skill 工具執行的 timeout（秒）
SKILL_TOOL_TIMEOUT = 60
# Skill agent 最大迭代次數（防止無限迴圈）
SKILL_MAX_ITERATIONS = 15
# Groq Free tier: 30 RPM → 每次 LLM 呼叫間隔至少 2 秒
SKILL_REQUEST_INTERVAL = 2.0
# 每 N 次 LLM 呼叫後強制冷卻
SKILL_COOLDOWN_EVERY = 14
SKILL_COOLDOWN_SECONDS = 60


@dataclass
class ValidationResult:
    status: str      # "ok" | "warning" | "failed" | "rate_limited"
    reason: str      # 中文說明
    suggestion: str  # LLM 建議的修復方向（failed 時才有意義）


def _is_rate_limit_error(e: Exception) -> bool:
    """偵測 LLM provider 的配額/速率錯誤。429 / RESOURCE_EXHAUSTED 都算。
    用於避免「驗證失敗 → fallback 再叫 LLM → 又 429」的連環燒配額。"""
    s = str(e)
    return ("429" in s or "RESOURCE_EXHAUSTED" in s or "quota" in s.lower()
            or "rate limit" in s.lower() or "rate_limit" in s.lower())


_llm = None
_llm_sig: Optional[str] = None


def _get_llm():
    global _llm, _llm_sig
    from settings import settings_signature
    from llm_factory import build_llm
    sig = settings_signature()
    if _llm is None or _llm_sig != sig:
        _llm = build_llm(temperature=0)
        _llm_sig = sig
    return _llm


# ── 檔案內容讀取 ──────────────────────────────────────────────────────────────

IMAGE_EXTS = {'.png', '.jpg', '.jpeg', '.gif', '.webp', '.bmp'}
STRUCTURED_EXTS = {'.csv', '.json', '.jsonl', '.xlsx', '.xls'}
TEXT_EXTS = {'.txt', '.log', '.md', '.html', '.xml', '.yaml', '.yml', '.py', '.sh', '.js', '.ts'}
MAX_TEXT_LINES = 50
MAX_CSV_ROWS = 10


def _read_file_content(path: Optional[str]) -> dict:
    """
    讀取輸出檔案，回傳結構化資訊供 LLM 分析。

    Returns:
        {
            "summary": str,       # 給 prompt 的文字摘要
            "image_b64": str|None # base64 圖片（僅圖檔）
            "image_mime": str|None
        }
    """
    result = {"summary": "", "image_b64": None, "image_mime": None}
    if not path:
        return result

    p = _resolve_user_path(path)
    if not p.exists():
        return result

    # 目錄：列出檔案清單
    if p.is_dir():
        files = sorted(p.iterdir())[:20]
        listing = "\n".join(f"  {'📁' if f.is_dir() else '📄'} {f.name} ({f.stat().st_size:,} bytes)" for f in files)
        result["summary"] = f"目錄內容（前 20 項）：\n{listing}"
        return result

    ext = p.suffix.lower()

    # 圖片檔 → base64
    if ext in IMAGE_EXTS:
        try:
            data = p.read_bytes()
            if len(data) <= 20 * 1024 * 1024:  # ≤ 20MB
                result["image_b64"] = base64.b64encode(data).decode()
                mime_map = {'.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
                           '.gif': 'image/gif', '.webp': 'image/webp', '.bmp': 'image/bmp'}
                result["image_mime"] = mime_map.get(ext, 'image/png')
                result["summary"] = f"圖片檔 {p.name}（{len(data):,} bytes），已附圖供視覺分析"
        except Exception as e:
            result["summary"] = f"圖片讀取失敗：{e}"
        return result

    # CSV
    if ext == '.csv':
        try:
            text = p.read_text(encoding='utf-8', errors='replace')
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if not rows:
                result["summary"] = "CSV 檔案為空"
                return result
            header = rows[0]
            data_rows = rows[1:]
            sample = data_rows[:MAX_CSV_ROWS]
            sample_str = "\n".join([",".join(r) for r in sample])
            result["summary"] = (
                f"CSV 檔案：{p.name}\n"
                f"欄位（{len(header)} 個）：{', '.join(header)}\n"
                f"資料行數：{len(data_rows)}\n"
                f"前 {min(len(sample), MAX_CSV_ROWS)} 行樣本：\n{sample_str}"
            )
        except Exception as e:
            result["summary"] = f"CSV 讀取失敗：{e}"
        return result

    # JSON / JSONL
    if ext in ('.json', '.jsonl'):
        try:
            text = p.read_text(encoding='utf-8', errors='replace')
            if ext == '.jsonl':
                lines = [l for l in text.strip().split('\n') if l.strip()]
                result["summary"] = (
                    f"JSONL 檔案：{p.name}，共 {len(lines)} 行\n"
                    f"前 {min(5, len(lines))} 行樣本：\n" +
                    "\n".join(lines[:5])
                )
            else:
                data = json.loads(text)
                if isinstance(data, list):
                    sample = json.dumps(data[:5], ensure_ascii=False, indent=2)
                    result["summary"] = f"JSON 陣列：{p.name}，共 {len(data)} 筆\n前 5 筆樣本：\n{sample}"
                elif isinstance(data, dict):
                    keys = list(data.keys())[:20]
                    result["summary"] = f"JSON 物件：{p.name}\n鍵（前 20 個）：{', '.join(keys)}\n內容預覽：\n{json.dumps(data, ensure_ascii=False, indent=2)[:1000]}"
                else:
                    result["summary"] = f"JSON 檔案：{p.name}\n內容：{text[:500]}"
        except Exception as e:
            result["summary"] = f"JSON 讀取失敗：{e}"
        return result

    # Excel
    if ext in ('.xlsx', '.xls'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
            sheets_info = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(max_row=MAX_CSV_ROWS + 1, values_only=True))
                if not rows:
                    sheets_info.append(f"  Sheet「{sheet_name}」：空")
                    continue
                header = [str(c) if c is not None else "" for c in rows[0]]
                data_rows = rows[1:]
                total_rows = ws.max_row - 1 if ws.max_row else 0
                sample_lines = []
                for r in data_rows[:MAX_CSV_ROWS]:
                    sample_lines.append(",".join(str(c) if c is not None else "" for c in r))
                sheets_info.append(
                    f"  Sheet「{sheet_name}」：{total_rows} 行，{len(header)} 欄\n"
                    f"    欄位：{', '.join(header)}\n"
                    f"    前 {len(sample_lines)} 行：\n    " + "\n    ".join(sample_lines)
                )
            wb.close()
            result["summary"] = f"Excel 檔案：{p.name}，共 {len(wb.sheetnames)} 個 Sheet\n" + "\n".join(sheets_info)
        except ImportError:
            result["summary"] = f"Excel 檔案：{p.name}（需安裝 openpyxl 才能讀取內容）"
        except Exception as e:
            result["summary"] = f"Excel 讀取失敗：{e}"
        return result

    # 一般文字檔（含未知副檔名）
    try:
        with open(p, 'r', encoding='utf-8', errors='replace') as f:
            lines = []
            for i, line in enumerate(f):
                if i >= MAX_TEXT_LINES:
                    break
                lines.append(line.rstrip())
        total_size = p.stat().st_size
        result["summary"] = (
            f"文字檔：{p.name}（{total_size:,} bytes）\n"
            f"前 {len(lines)} 行：\n" + "\n".join(lines)
        )
    except Exception:
        result["summary"] = f"檔案 {p.name} 無法以文字方式讀取"

    return result


# ── 主驗證函式 ─────────────────────────────────────────────────────────────────

async def validate_step(
    step_name: str,
    command: str,
    exit_code: int,
    stdout: str,
    stderr: str,
    output_path: Optional[str],
    output_expect: Optional[str],
    logger: logging.Logger,
) -> ValidationResult:
    """
    使用 LLM 語意分析執行結果，回傳結構化驗證結論。

    LLM 會考量：
    - exit code 與其含意
    - stdout/stderr 的語意（區分警告與錯誤）
    - 輸出檔案是否存在、大小是否合理
    - 輸出檔案內容（文字前 50 行 / CSV 結構 / Excel 摘要）
    - 圖片檔案以視覺方式驗證
    - 是否符合 expect 描述的期望
    """
    # 收集輸出檔案資訊
    file_info = _check_output_file(output_path)
    file_content = _read_file_content(output_path)

    # 截取重要片段（節省 token）
    stdout_tail = stdout[-1000:] if len(stdout) > 1000 else stdout
    stderr_tail = stderr[-500:] if len(stderr) > 500 else stderr

    prompt_text = f"""你是一個精確的 pipeline 步驟驗證器。
分析以下執行結果，判斷步驟是否成功。

【步驟資訊】
名稱：{step_name}
命令：{command}
Exit Code：{exit_code}
預期輸出描述：{output_expect or "無特定要求"}
輸出路徑：{output_path or "無"}
檔案狀態：{file_info}

【stdout（最後部分）】
```
{stdout_tail or "（無輸出）"}
```

【stderr（最後部分）】
```
{stderr_tail or "（無輸出）"}
```"""

    # 加入檔案內容摘要
    if file_content["summary"]:
        prompt_text += f"""

【輸出檔案內容】
{file_content["summary"]}"""

    # 如果是圖片，加入視覺分析提示
    if file_content["image_b64"]:
        prompt_text += """

【圖片分析】
已附上輸出的圖片檔案，請以視覺方式分析圖片內容是否符合預期描述。
檢查圖片是否正常渲染、內容是否完整、是否符合期望。"""

    prompt_text += """

請只回傳以下 JSON，不要加任何其他文字：
{
  "status": "ok",
  "reason": "一句話說明判斷結果",
  "suggestion": "如果 failed，給出修復建議；ok 時留空字串"
}

【判斷規則】
- "ok"：步驟成功，exit code 0，輸出符合預期（若有）
- "warning"：步驟完成但有非致命問題（如 deprecation warning、部分資料遺失），建議人工確認
- "failed"：步驟失敗，需要介入（exit code 非 0 且 stderr 有真實錯誤、Exception、缺少必要輸出檔案等）

注意：Python DeprecationWarning、UserWarning 不代表失敗；只有真正的 Exception / Error / 致命問題才判為 failed。"""

    try:
        llm = _get_llm()

        # 構建 message content（支援圖片 vision）
        if file_content["image_b64"]:
            content = [
                {"type": "text", "text": prompt_text},
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:{file_content['image_mime']};base64,{file_content['image_b64']}"
                    },
                },
            ]
        else:
            content = prompt_text

        from llm_factory import invoke_with_streaming
        raw = (await invoke_with_streaming(
            llm,
            [
                SystemMessage(content="你是一個精確的 pipeline 驗證器，只回傳 JSON 格式。"),
                HumanMessage(content=content),
            ],
            label=f"validator:{step_name}",
            timeout=300.0,
            logger=logger,
        )).strip()
        # 去除 markdown code block（如果有）
        if "```" in raw:
            parts = raw.split("```")
            raw = parts[1].strip()
            if raw.startswith("json"):
                raw = raw[4:].strip()

        data = json.loads(raw)
        result = ValidationResult(
            status=data.get("status", "failed"),
            reason=data.get("reason", ""),
            suggestion=data.get("suggestion", ""),
        )
        logger.info(f"[{step_name}] 驗證：{result.status} — {result.reason}")
        return result

    except Exception as e:
        # 429 / RESOURCE_EXHAUSTED：不退 fallback，直接回 rate_limited 給 runner，
        # 避免下一條 fallback 路徑又叫一次 LLM 再燒一次配額
        if _is_rate_limit_error(e):
            logger.error(f"[{step_name}] LLM 配額/速率受限（429）— 不退 fallback 避免燒光配額：{str(e)[:200]}")
            return ValidationResult(
                status="rate_limited",
                reason=f"LLM provider 配額用盡或速率受限（429）：{str(e)[:300]}",
                suggestion="等配額重置（通常每分鐘 / 每天）或在 Settings 切換 provider（Groq / OpenRouter / Ollama 本地）",
            )
        logger.error(f"[{step_name}] LLM 驗證失敗：{e}，退回 exit code 判斷")
        # Fallback：純 exit code 判斷
        if exit_code == 0:
            return ValidationResult(
                status="ok",
                reason=f"Exit code 0（LLM 驗證服務暫時不可用：{e}）",
                suggestion="",
            )
        return ValidationResult(
            status="failed",
            reason=f"Exit code {exit_code}（LLM 驗證服務暫時不可用：{e}）",
            suggestion="請檢查 log 檔取得詳細錯誤訊息",
        )


def _resolve_user_path(path: str) -> Path:
    """統一處理使用者可能給的三種路徑：
    - 絕對路徑 → 直接用
    - `~/xxx` → 展開到使用者家目錄
    - 相對路徑 → 以**專案根目錄**為基準（非 backend cwd），跟 runner 邏輯一致
    """
    p = Path(path).expanduser()
    if not p.is_absolute():
        _PROJ_ROOT = Path(__file__).parent.parent.parent.absolute()
        p = _PROJ_ROOT / p
    return p


def _check_output_file(path: Optional[str]) -> str:
    """取得輸出檔案或目錄的基本資訊"""
    if not path:
        return "無需檢查"
    p = _resolve_user_path(path)
    if not p.exists():
        return "❌ 路徑不存在"
    if p.is_dir():
        files = list(p.iterdir())
        if not files:
            return "⚠ 目錄存在但為空"
        total = sum(f.stat().st_size for f in files if f.is_file())
        return f"✅ 目錄存在，共 {len(files)} 個檔案，總大小：{total:,} bytes"
    size = p.stat().st_size
    if size == 0:
        return "⚠ 檔案存在但為空（0 bytes）"
    return f"✅ 檔案存在，大小：{size:,} bytes"


# ── Skill 模式：ReAct Agent 驗證 ──────────────────────────────────────────────

# 危險命令黑名單（防止 LLM 生成危險操作）
_DANGEROUS_COMMANDS = {'rm', 'rmdir', 'del', 'format', 'mkfs', 'dd', 'kill', 'shutdown', 'reboot'}


def _run_python_sync(code: str) -> str:
    """在 subprocess 中執行 Python 程式碼，回傳 stdout + stderr。"""
    # 截斷混入程式碼中的 <tool> 標籤
    tool_tag_pos = code.find('<tool>')
    if tool_tag_pos > 0:
        code = code[:tool_tag_pos].rstrip()
    # 注入 done/view_image/read_file 的 no-op stub，避免 LLM 把工具名當 Python 函式呼叫而崩潰
    preamble = (
        "# -*- coding: utf-8 -*-\n"
        "import warnings\n"
        "warnings.filterwarnings('ignore')\n"
        "def done(*args, **kwargs):\n"
        "    print('[info] done() is a tool, not a Python function - ignored')\n"
        "def view_image(*args, **kwargs):\n"
        "    print('[info] view_image() is a tool, not a Python function - ignored')\n"
        "def read_file(*args, **kwargs):\n"
        "    print('[info] read_file() is a tool, not a Python function - ignored')\n"
    )
    code = preamble + code
    # wsl_docker 模式：路由到沙盒；沙盒不可用 / host 模式時 fallback 到下面 subprocess
    try:
        from pipeline.executor import _try_sandbox_exec
        sandbox_out = _try_sandbox_exec("run_python", code, None, "", None)
        if sandbox_out is not None:
            return sandbox_out
    except Exception:
        pass  # 沙盒模組壞了也 fallback 到 host subprocess
    try:
        # UTF-8 寫檔（見 executor.py 同樣 fix 的註解）
        with tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False, encoding='utf-8') as f:
            f.write(code)
            tmp_path = f.name
        from pipeline.executor import _SKILL_PYTHON, _clean_env
        result = subprocess.run(
            [_SKILL_PYTHON, tmp_path],
            capture_output=True, text=True,
            timeout=SKILL_TOOL_TIMEOUT,
            env=_clean_env(),  # 套用 PYTHONIOENCODING=utf-8 防中文 print 爆炸
        )
        Path(tmp_path).unlink(missing_ok=True)
        output = ""
        if result.stdout:
            output += result.stdout
        if result.stderr:
            tag = "stderr" if result.returncode != 0 else "warnings"
            output += f"\n[{tag}]\n{result.stderr}"
        if result.returncode != 0:
            output += f"\n[exit code: {result.returncode}]"
        elif not result.stdout:
            output += "\n[執行成功，程式無 stdout 輸出]"
        return output.strip() or "(無輸出)"
    except subprocess.TimeoutExpired:
        Path(tmp_path).unlink(missing_ok=True)
        return f"[錯誤] Python 執行超時（>{SKILL_TOOL_TIMEOUT}秒）"
    except Exception as e:
        return f"[錯誤] Python 執行失敗：{e}"


def _run_shell_sync(cmd: str) -> str:
    """執行 shell 命令，回傳輸出。會過濾危險命令。
    wsl_docker 模式下透過 executor._try_sandbox_exec 路由到容器；
    否則走 host subprocess（原行為）。
    """
    first_word = cmd.strip().split()[0] if cmd.strip() else ""
    if first_word in _DANGEROUS_COMMANDS:
        return f"[拒絕] 命令 '{first_word}' 被安全策略封鎖"
    # 先試沙盒（如果 settings.skill_sandbox_mode='wsl_docker'）
    try:
        from pipeline.executor import _try_sandbox_exec
        sandbox_out = _try_sandbox_exec("run_shell", cmd, None, "", None)
        if sandbox_out is not None:
            return sandbox_out
    except Exception:
        pass  # 沙盒模組問題 → 繼續走 host fallback
    # 統一 python interpreter（與 executor._skill_run_shell 一致）
    from pipeline.executor import _rewrite_python_cmd
    cmd = _rewrite_python_cmd(cmd)
    try:
        result = subprocess.run(
            cmd, shell=True,
            capture_output=True, text=True,
            timeout=SKILL_TOOL_TIMEOUT,
        )
        output = ""
        if result.stdout:
            output += result.stdout
        if result.stderr:
            output += f"\n[stderr]\n{result.stderr}"
        return output.strip()[:5000] or "(無輸出)"
    except subprocess.TimeoutExpired:
        return f"[錯誤] 命令執行超時（>{SKILL_TOOL_TIMEOUT}秒）"
    except Exception as e:
        return f"[錯誤] 命令執行失敗：{e}"


def _read_file_sync(path: str, max_lines: int = 100) -> str:
    """讀取檔案內容（最多 max_lines 行）。"""
    try:
        # 清理 LLM 常見的錯誤格式：read_file("path"), 引號, 空白
        cleaned = path.strip()
        import re as _re
        m = _re.match(r'read_file\(["\']?(.+?)["\']?\)\s*$', cleaned)
        if m:
            cleaned = m.group(1)
        cleaned = cleaned.strip().strip('"').strip("'")
        # 沙盒路徑 → Windows 路徑（同 _view_image_sync 同份補丁）
        m_wsl = _re.match(r"^/mnt/([a-z])/(.*)$", cleaned)
        if m_wsl:
            cleaned = f"{m_wsl.group(1).upper()}:\\{m_wsl.group(2).replace('/', chr(92))}"
        p = Path(cleaned).expanduser()
        if not p.exists():
            return f"[錯誤] 檔案不存在：{path}（解析後：{p}）"
        if p.is_dir():
            files = sorted(p.iterdir())[:30]
            listing = "\n".join(f"  {'📁' if f.is_dir() else '📄'} {f.name} ({f.stat().st_size:,} B)" for f in files)
            return f"目錄內容：\n{listing}"
        # 偵測二進制檔案，避免汙染 LLM context
        binary_exts = {'.xlsx', '.xls', '.docx', '.pptx', '.pdf', '.png', '.jpg', '.jpeg', '.gif', '.zip', '.gz', '.tar', '.pkl', '.npy', '.parquet'}
        if p.suffix.lower() in binary_exts:
            size = p.stat().st_size
            return (f"[提示] {p.name} 是二進制檔案（{size:,} bytes），無法用 read_file 讀取。\n"
                    f"請改用 run_python 搭配適當的套件讀取：\n"
                    f"- .xlsx/.xls → pandas.read_excel() 或 openpyxl\n"
                    f"- .docx → python-docx\n"
                    f"- .png/.jpg → PIL 或 view_image 工具\n"
                    f"- .pdf → PyPDF2")
        if p.stat().st_size > 10 * 1024 * 1024:
            return f"[警告] 檔案過大（{p.stat().st_size:,} bytes），只讀前 {max_lines} 行"
        with open(p, 'r', encoding='utf-8', errors='replace') as f:
            lines = []
            for i, line in enumerate(f):
                if i >= max_lines:
                    lines.append(f"... (已截斷，共超過 {max_lines} 行)")
                    break
                lines.append(line.rstrip())
        return "\n".join(lines) or "(空檔案)"
    except Exception as e:
        return f"[錯誤] 讀取失敗：{e}"


def _sanitize_code(code: str) -> str:
    """清除混入程式碼中的 LLM 解釋文字（非 Python/Shell 語法的行）。"""
    lines = code.split('\n')
    code_pattern = re.compile(
        r'^(\s*(import |from |def |class |if |for |while |with |try:|except |'
        r'return |print|#|[a-zA-Z_]\w*\s*[=(]|plt\.|df\.|pd\.|np\.|sns\.|'
        r'\[|{|}|\]|\)|"|\'|$))'
    )
    start_idx = 0
    for i, line in enumerate(lines):
        stripped = line.strip()
        if not stripped:
            continue
        if code_pattern.match(stripped):
            start_idx = i
            break
    result = []
    for line in lines[start_idx:]:
        stripped = line.strip()
        if not stripped:
            result.append(line)
            continue
        first_char = stripped[0]
        if ord(first_char) > 0x2E00 and not stripped.startswith('#') and not stripped.startswith(("'", '"')):
            continue
        result.append(line)
    return '\n'.join(result).strip()


def _parse_tool_calls(text: str) -> list[dict]:
    """
    從 LLM 回覆中解析工具呼叫。
    支援：標準 <input> 標籤、code block 包裹、無標籤直接跟內容。
    關鍵：確保 run_python 只提取程式碼，不混入 LLM 解釋文字。
    """
    import re
    calls = []

    # Step 1：標準 <tool>...</tool> <input>...</input>
    pattern_std = re.compile(r'<tool>(.*?)</tool>\s*<input>(.*?)</input>', re.DOTALL)
    for m in pattern_std.finditer(text):
        calls.append({"tool": m.group(1).strip(), "input": m.group(2).strip()})
    if calls:
        return calls

    # Step 2：找所有 code blocks，再找離 <tool> 最近的那個
    code_blocks = list(re.finditer(r'```(?:python|json|bash|sh)?\s*\n(.*?)```', text, re.DOTALL))
    tool_tags = list(re.finditer(r'<tool>(.*?)</tool>', text))

    for tag in tool_tags:
        tool_name = tag.group(1).strip()
        tag_start = tag.start()
        tag_end = tag.end()

        # 先找 tag 之後最近的 code block
        best_block = None
        for block in code_blocks:
            if block.start() >= tag_end:
                best_block = block
                break

        # 如果 tag 之後沒有 code block，往前找最近的（LLM 先放 code 再放 tag）
        if not best_block:
            for block in reversed(code_blocks):
                if block.end() <= tag_start:
                    best_block = block
                    break

        if best_block:
            content = best_block.group(1).strip()
            if tool_name in ('run_python', 'run_shell'):
                content = _sanitize_code(content)
            if content and len(content) > 2:
                calls.append({"tool": tool_name, "input": content})
                return calls

    # Step 3：done 工具 — 找 JSON
    done_match = re.search(r'<tool>done</tool>', text)
    if done_match:
        after_done = text[done_match.end():]
        json_match = re.search(r'\{.*?\}', after_done, re.DOTALL)
        if json_match:
            return [{"tool": "done", "input": json_match.group(0).strip()}]

    # Step 4：沒有 <tool> 標籤但有 code block
    if not tool_tags and code_blocks:
        content = code_blocks[-1].group(1).strip()
        if content.startswith('{') and ('success' in content or 'status' in content):
            return [{"tool": "done", "input": content}]

    # Step 5：fallback
    cleaned = re.sub(r'```(?:python|json|bash|sh)?\s*\n?', '', text)
    cleaned = cleaned.replace('```', '')
    pattern_raw = re.compile(r'<tool>(.*?)</tool>\s*(.+?)(?=<tool>|$)', re.DOTALL)
    for m in pattern_raw.finditer(cleaned):
        tool_name = m.group(1).strip()
        content = m.group(2).strip()
        if tool_name in ('run_python', 'run_shell'):
            content = _sanitize_code(content)
        if content and len(content) > 2:
            calls.append({"tool": tool_name, "input": content})
            break
    return calls


IMAGE_EXTS_SKILL = {'.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
                    '.gif': 'image/gif', '.webp': 'image/webp', '.bmp': 'image/bmp'}


def _view_image_sync(path: str) -> dict:
    """
    讀取圖片並回傳 base64 資料。
    回傳 {"text": 描述, "image_b64": str|None, "image_mime": str|None}
    """
    try:
        cleaned = path.strip().strip('"').strip("'")
        # 沙盒路徑 → Windows 路徑（V3 view_image bug：LLM 跑沙盒給 /mnt/c/... 結果讀不到）
        import re as _re
        m = _re.match(r"^/mnt/([a-z])/(.*)$", cleaned)
        if m:
            cleaned = f"{m.group(1).upper()}:\\{m.group(2).replace('/', chr(92))}"
        p = Path(cleaned).expanduser()
        if not p.exists():
            return {"text": f"[錯誤] 圖片不存在：{path}（解析後：{p}）", "image_b64": None, "image_mime": None}
        ext = p.suffix.lower()
        if ext not in IMAGE_EXTS_SKILL:
            return {"text": f"[錯誤] 不支援的圖片格式：{ext}，支援 {list(IMAGE_EXTS_SKILL.keys())}", "image_b64": None, "image_mime": None}
        data = p.read_bytes()
        if len(data) > 20 * 1024 * 1024:
            return {"text": f"[錯誤] 圖片過大（{len(data):,} bytes，上限 20MB）", "image_b64": None, "image_mime": None}
        b64 = base64.b64encode(data).decode()
        mime = IMAGE_EXTS_SKILL[ext]
        return {"text": f"圖片 {p.name}（{len(data):,} bytes），已載入供視覺分析", "image_b64": b64, "image_mime": mime}
    except Exception as e:
        return {"text": f"[錯誤] 圖片讀取失敗：{e}", "image_b64": None, "image_mime": None}


def _execute_tool(tool_name: str, tool_input: str) -> str:
    """執行單一工具呼叫（非圖片工具）。"""
    if tool_name == "run_python":
        return _run_python_sync(tool_input)
    elif tool_name == "run_shell":
        return _run_shell_sync(tool_input)
    elif tool_name == "read_file":
        return _read_file_sync(tool_input.strip())
    elif tool_name == "done":
        return "__DONE__"
    elif tool_name == "view_image":
        return "__VIEW_IMAGE__"  # 特殊標記，在 agent loop 中處理
    else:
        return f"[錯誤] 未知工具：{tool_name}"


async def validate_step_with_skill(
    step_name: str,
    command: str,
    exit_code: int,
    stdout: str,
    stderr: str,
    output_path: Optional[str],
    output_expect: Optional[str],
    logger: logging.Logger,
) -> ValidationResult:
    """
    Skill 模式驗證：LLM 作為 ReAct agent，可主動執行程式碼來驗證步驟結果。

    工具：
    - run_python(code): 執行 Python 程式碼
    - run_shell(cmd): 執行 Shell 命令
    - read_file(path): 讀取檔案內容
    - done(json): 結束驗證並回傳結果
    """
    stdout_tail = stdout[-1500:] if len(stdout) > 1500 else stdout
    stderr_tail = stderr[-500:] if len(stderr) > 500 else stderr

    system_prompt = """你是一個 pipeline 步驟的 Skill 驗證 agent。
你可以主動執行程式碼來驗證步驟的輸出是否正確，而不只是被動地閱讀文字。

你有以下工具可用：

1. run_python — 執行 Python 程式碼
   用法：<tool>run_python</tool>
   <input>
   import pandas as pd
   df = pd.read_csv("/path/to/file.csv")
   print(f"行數：{len(df)}")
   print(f"欄位：{list(df.columns)}")
   </input>

2. run_shell — 執行系統命令
   用法：<tool>run_shell</tool>
   <input>wc -l /path/to/file.csv</input>
   注意：盡量用 run_python 代替 run_shell，因為 Python 是跨平台的。

3. read_file — 讀取檔案內容
   用法：<tool>read_file</tool>
   <input>/path/to/file.csv</input>

4. view_image — 查看圖片（視覺分析，支援 png/jpg/gif/webp）
   用法：<tool>view_image</tool>
   <input>/path/to/chart.png</input>
   系統會將圖片顯示給你，你可以用視覺判斷圖片內容是否正確。
   適用場景：驗證圖表是否有標題、座標軸、資料是否合理、圖片是否正常渲染等。

5. done — 結束驗證，回傳最終結果（必須是 JSON）
   用法：<tool>done</tool>
   <input>{"status": "ok", "reason": "說明", "suggestion": ""}</input>

【可用 Python 套件】
標準庫：csv, json, random, os, pathlib, re, math, datetime, io, collections
資料處理：pandas, numpy, openpyxl, xlrd, tabulate
文件處理：python-docx (docx), python-pptx (pptx), PyPDF2, reportlab, jinja2
網頁/爬蟲：requests, beautifulsoup4 (bs4), lxml
圖表繪製：matplotlib, seaborn, plotly
圖片處理：Pillow (PIL)
其他：faker, pyyaml, chardet

【matplotlib 繪圖注意事項】
- 使用 matplotlib.pyplot 時，務必在最前面加 `import matplotlib; matplotlib.use('Agg')` 以避免 GUI 問題
- boxplot 的 `labels` 參數已在新版棄用，請改用 `tick_labels`
- 繪製分組箱形圖時，需要先將資料按分組欄位 pivot/reshape，再分別傳入各組資料
- 中文顯示：macOS 使用 'PingFang HK'；Windows 使用 'Microsoft JhengHei' 或 'SimHei'
- 繪圖完成後務必呼叫 `plt.savefig(路徑, dpi=150, bbox_inches='tight')` 並 `plt.close()`

【重要規則】
- **路徑處理：一律使用 `pathlib.Path` 或 `os.path.join` 組合路徑，不要用字串拼接 `/`**
- **只使用上方列出的已安裝套件，不要安裝新套件**
- **絕對不要執行 sudo、pip install、apt 等安裝命令**
- 根據「預期輸出描述」主動驗證，不要只看 exit code
- **如果輸出路徑是圖片檔（.png/.jpg 等），一定要使用 view_image 工具查看圖片內容再做判斷**
- 可以多次呼叫工具，逐步分析
- 每次只呼叫一個工具
- 最後一定要呼叫 done 工具回傳結論
- status 只能是 "ok"、"warning"、"failed" 三者之一
- reason 和 suggestion 用中文"""

    user_prompt = f"""請驗證以下 pipeline 步驟的執行結果：

【步驟資訊】
名稱：{step_name}
命令：{command}
Exit Code：{exit_code}
預期輸出描述：{output_expect or "無特定要求"}
輸出路徑：{output_path or "無"}

【stdout（最後部分）】
```
{stdout_tail or "（無輸出）"}
```

【stderr（最後部分）】
```
{stderr_tail or "（無輸出）"}
```

請使用工具主動驗證輸出是否符合預期。開始吧。"""

    try:
        llm = _get_llm()
        messages = [
            SystemMessage(content=system_prompt),
            HumanMessage(content=user_prompt),
        ]

        for iteration in range(SKILL_MAX_ITERATIONS):
            logger.info(f"[{step_name}] Skill agent 迭代 {iteration + 1}/{SKILL_MAX_ITERATIONS}")

            # 冷卻機制
            if iteration > 0 and iteration % SKILL_COOLDOWN_EVERY == 0:
                logger.info(f"[{step_name}] ⏸ 達到 {SKILL_COOLDOWN_EVERY} 次呼叫，冷卻 {SKILL_COOLDOWN_SECONDS} 秒...")
                await asyncio.sleep(SKILL_COOLDOWN_SECONDS)

            if iteration > 0:
                await asyncio.sleep(SKILL_REQUEST_INTERVAL)

            from llm_factory import invoke_with_streaming
            reply = (await invoke_with_streaming(
                llm, messages, label=f"validator:{step_name}", timeout=180.0, logger=logger
            )).strip()
            _vp = reply if len(reply) <= 4000 else reply[:4000] + f"...[已截斷，完整長度 {len(reply)} 字]"
            logger.debug(f"[{step_name}] Agent 回覆：\n{_vp}")

            # 解析工具呼叫
            tool_calls = _parse_tool_calls(reply)

            if not tool_calls:
                # 嘗試直接解析為 JSON（LLM 可能直接回傳結果）
                try:
                    raw = reply
                    if "```" in raw:
                        parts = raw.split("```")
                        raw = parts[1].strip()
                        if raw.startswith("json"):
                            raw = raw[4:].strip()
                    data = json.loads(raw)
                    if "status" in data:
                        result = ValidationResult(
                            status=data.get("status", "failed"),
                            reason=data.get("reason", ""),
                            suggestion=data.get("suggestion", ""),
                        )
                        logger.info(f"[{step_name}] Skill 驗證：{result.status} — {result.reason}")
                        return result
                except (json.JSONDecodeError, IndexError):
                    pass
                # 無法解析，加入提示讓 agent 繼續
                messages.append(HumanMessage(content=reply))
                messages.append(HumanMessage(content="請使用工具來驗證，或呼叫 done 工具回傳最終結論。"))
                continue

            # 執行第一個工具
            call = tool_calls[0]
            tool_name = call["tool"]
            tool_input = call["input"]

            # done 工具 → 結束
            if tool_name == "done":
                try:
                    data = json.loads(tool_input)
                    result = ValidationResult(
                        status=data.get("status", "failed"),
                        reason=data.get("reason", ""),
                        suggestion=data.get("suggestion", ""),
                    )
                    logger.info(f"[{step_name}] Skill 驗證完成：{result.status} — {result.reason}")
                    return result
                except json.JSONDecodeError:
                    messages.append(HumanMessage(content=reply))
                    messages.append(HumanMessage(content=f"[系統] done 的 input 必須是有效 JSON，請重試。"))
                    continue

            # 執行工具
            logger.info(f"[{step_name}] 執行工具 {tool_name}")

            # view_image 特殊處理：注入多模態訊息
            if tool_name == "view_image":
                img_data = await asyncio.get_event_loop().run_in_executor(
                    None, _view_image_sync, tool_input
                )
                logger.debug(f"[{step_name}] view_image：{img_data['text']}")
                messages.append(HumanMessage(content=reply))
                if img_data["image_b64"]:
                    messages.append(HumanMessage(content=[
                        {"type": "text", "text": f"[工具結果 — view_image]\n{img_data['text']}\n請仔細觀察圖片內容，判斷是否符合預期。"},
                        {"type": "image_url", "image_url": {
                            "url": f"data:{img_data['image_mime']};base64,{img_data['image_b64']}"
                        }},
                    ]))
                else:
                    messages.append(HumanMessage(content=f"[工具結果 — view_image]\n{img_data['text']}"))
                continue

            tool_result = await asyncio.get_event_loop().run_in_executor(
                None, _execute_tool, tool_name, tool_input
            )
            _vt = tool_result if len(tool_result) <= 3000 else tool_result[:3000] + f"...[已截斷，完整長度 {len(tool_result)} 字]"
            logger.debug(f"[{step_name}] 工具結果：\n{_vt}")

            # 加入對話歷史
            messages.append(HumanMessage(content=reply))
            messages.append(HumanMessage(content=f"[工具結果 — {tool_name}]\n{tool_result}"))

        # 超過最大迭代次數
        logger.warning(f"[{step_name}] Skill agent 達到最大迭代次數")
        return ValidationResult(
            status="warning",
            reason=f"Skill agent 在 {SKILL_MAX_ITERATIONS} 次迭代內未完成驗證",
            suggestion="建議手動檢查輸出結果",
        )

    except Exception as e:
        # 429 / RESOURCE_EXHAUSTED：直接回 rate_limited，不退 validate_step（會再 429 一次）
        if _is_rate_limit_error(e):
            logger.error(f"[{step_name}] Skill 驗證 LLM 配額/速率受限（429）— 不退一般驗證避免燒光配額：{str(e)[:200]}")
            return ValidationResult(
                status="rate_limited",
                reason=f"LLM provider 配額用盡或速率受限（429）：{str(e)[:300]}",
                suggestion="等配額重置或在 Settings 切換 provider（Groq / OpenRouter / Ollama 本地）",
            )
        logger.error(f"[{step_name}] Skill 驗證失敗：{e}，退回一般驗證")
        # Fallback to standard validation
        return await validate_step(
            step_name=step_name,
            command=command,
            exit_code=exit_code,
            stdout=stdout,
            stderr=stderr,
            output_path=output_path,
            output_expect=output_expect,
            logger=logger,
        )
