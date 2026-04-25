"""
【工具 D】Q1 財務報表產生器
製作者：財務長辦公室 / 管理報表組 (CFO Office)

用途：讀取財務分析結果，產生格式化的管理報表 Excel，
      包含封面頁、執行摘要、各維度明細，
      套用樣式（標題顏色、粗體、數字格式、邊框）。

輸入：~/ai_output/finance/financial_summary.xlsx
輸出：~/ai_output/finance/Q1_financial_report.xlsx
"""
import os
import sys
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from pathlib import Path

# 載入 .env 設定
load_dotenv(Path(__file__).parent.parent.parent / "backend" / ".env")

# ── 路徑設定 ────────────────────────────────────────────────────────────────
def get_paths():
    # 1. 優先使用主程式動態注入的路徑 (包含工作流名稱)
    env_run_dir = os.getenv("PIPELINE_OUTPUT_DIR")
    if env_run_dir:
        return (
            os.path.join(env_run_dir, "financial_summary.xlsx"),
            os.path.join(env_run_dir, "Q1_financial_report.xlsx")
        )

    # 2. 次優先從環境變數讀取全局輸出根目錄
    base_path = os.getenv("OUTPUT_BASE_PATH")
    if base_path:
        if not os.path.isabs(base_path):
             base_path = os.path.join(Path(__file__).parent.parent.parent, base_path)
        return (
            os.path.join(base_path, "finance", "financial_summary.xlsx"),
            os.path.join(base_path, "finance", "Q1_financial_report.xlsx")
        )
    
    # 3. 預設：專案根目錄/ai_output/finance
    project_root = Path(__file__).parent.parent.parent
    return (
        os.path.join(project_root, "ai_output", "finance", "financial_summary.xlsx"),
        os.path.join(project_root, "ai_output", "finance", "Q1_financial_report.xlsx")
    )

INPUT, OUTPUT = get_paths()

if not os.path.exists(INPUT):
    print(f"[ERROR] 找不到輸入檔案：{INPUT}")
    print("        請先執行 stage3_analyze_finance.py")
    sys.exit(1)

# ── 讀取分析結果 ─────────────────────────────────────────────────────────────
kpi_df     = pd.read_excel(INPUT, sheet_name="KPI 總覽")
dept_df    = pd.read_excel(INPUT, sheet_name="部門別收支")
monthly_df = pd.read_excel(INPUT, sheet_name="月份別趨勢")
cat_df     = pd.read_excel(INPUT, sheet_name="費用類別排名")
rev_df     = pd.read_excel(INPUT, sheet_name="收入來源結構")

# KPI 快捷變數
def kpi(name):
    row = kpi_df[kpi_df["指標"] == name]
    return row["金額 (USD)"].values[0] if len(row) else 0

total_rev  = kpi("Q1 總收入")
total_exp  = kpi("Q1 總支出")
net_income = kpi("Q1 淨利")
exp_ratio  = kpi("費用率")

# ── 樣式定義 ─────────────────────────────────────────────────────────────────
NAVY   = PatternFill("solid", fgColor="1F3864")
BLUE   = PatternFill("solid", fgColor="2E75B6")
LTBLUE = PatternFill("solid", fgColor="BDD7EE")
GREEN  = PatternFill("solid", fgColor="E2EFDA")
WHITE  = PatternFill("solid", fgColor="FFFFFF")

def h_font(size=12, bold=True, color="FFFFFF"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(bold=False):
    return Font(name="Calibri", size=11, bold=bold, color="000000")

def thin_border():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def right_align():
    return Alignment(horizontal="right", vertical="center")

def style_header_row(ws, row, fill, font=None):
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font or h_font()
        cell.alignment = center()
        cell.border = thin_border()

def auto_width(ws, min_w=12, max_w=40):
    for col in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

# ── 建立報表 ─────────────────────────────────────────────────────────────────
with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:

    # ── 封面頁 ──────────────────────────────────────────────────────────────
    cover_data = pd.DataFrame([
        ["", ""],
        ["公司名稱",    "TechCorp International Ltd."],
        ["報表標題",    "2024 年 Q1 財務績效報告"],
        ["報表期間",    "2024-01-01 ～ 2024-03-31"],
        ["製表日期",    datetime.today().strftime("%Y-%m-%d")],
        ["機密等級",    "CONFIDENTIAL – 限管理階層"],
        ["", ""],
        ["核心指標",    ""],
        ["Q1 總收入",   f"USD {total_rev:,.2f}"],
        ["Q1 總支出",   f"USD {total_exp:,.2f}"],
        ["Q1 淨利",     f"USD {net_income:,.2f}"],
        ["費用率",      f"{exp_ratio:.1f}%"],
    ])
    cover_data.to_excel(writer, sheet_name="封面", index=False, header=False)
    ws = writer.sheets["封面"]
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 38
    for r in ws.iter_rows():
        for c in r:
            c.alignment = Alignment(vertical="center")
            c.font = body_font()
    for row_idx, label in [(2,"公司名稱"),(3,"報表標題"),(4,"報表期間"),(5,"製表日期"),(6,"機密等級")]:
        ws.cell(row_idx, 1).font = body_font(bold=True)
    for row_idx in [9, 10, 11, 12]:
        ws.cell(row_idx, 1).font = body_font(bold=True)
        ws.cell(row_idx, 2).font = Font(name="Calibri", size=12, bold=True, color="1F3864")
    ws.cell(8, 1).fill = NAVY
    ws.cell(8, 1).font = h_font(14)
    ws.cell(8, 1).value = "核心指標摘要"
    ws.cell(8, 2).fill = NAVY

    # ── 執行摘要 ────────────────────────────────────────────────────────────
    kpi_df.to_excel(writer, sheet_name="執行摘要", index=False)
    ws = writer.sheets["執行摘要"]
    style_header_row(ws, 1, NAVY)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font()
            cell.border = thin_border()
            cell.fill = LTBLUE if cell.row % 2 == 0 else WHITE
    auto_width(ws)

    # ── 部門別收支 ──────────────────────────────────────────────────────────
    dept_df.to_excel(writer, sheet_name="部門別收支", index=False)
    ws = writer.sheets["部門別收支"]
    style_header_row(ws, 1, BLUE)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font()
            cell.border = thin_border()
            cell.fill = GREEN if cell.row % 2 == 0 else WHITE
    auto_width(ws)

    # ── 月份別趨勢 ──────────────────────────────────────────────────────────
    monthly_df.to_excel(writer, sheet_name="月份別趨勢", index=False)
    ws = writer.sheets["月份別趨勢"]
    style_header_row(ws, 1, BLUE)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font()
            cell.border = thin_border()
    auto_width(ws)

    # ── 費用類別排名 ────────────────────────────────────────────────────────
    cat_df.to_excel(writer, sheet_name="費用類別排名", index=False)
    ws = writer.sheets["費用類別排名"]
    style_header_row(ws, 1, NAVY)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font()
            cell.border = thin_border()
            cell.fill = LTBLUE if cell.row % 2 == 0 else WHITE
    auto_width(ws)

    # ── 收入來源結構 ────────────────────────────────────────────────────────
    rev_df.to_excel(writer, sheet_name="收入來源結構", index=False)
    ws = writer.sheets["收入來源結構"]
    style_header_row(ws, 1, NAVY)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font()
            cell.border = thin_border()
            cell.fill = GREEN if cell.row % 2 == 0 else WHITE
    auto_width(ws)

print("=" * 55)
print("Stage 4：Q1 財務報表產生完成")
print("=" * 55)
print(f"  工作表      : 封面 / 執行摘要 / 部門別收支 / 月份別趨勢")
print(f"              : 費用類別排名 / 收入來源結構")
print(f"  Q1 淨利     : USD {net_income:,.2f}")
print(f"  輸出路徑    : {OUTPUT}")
print("  ✅ 報表已套用格式樣式，可直接提交管理層審閱")
